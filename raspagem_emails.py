import requests
from bs4 import BeautifulSoup
import openpyxl
import os

def buscar_emails_por_cnpj(caminho_excel, site_origem):
    """
    Realiza a busca de e-mails no site 'cnpja.com/office/{CNPJ}' com base nos CNPJs encontrados no Excel.
    
    Parâmetros:
        caminho_excel (str): Caminho do arquivo Excel com os CNPJs.
        site_origem (str): Nome do site que originou o arquivo (para identificação).
    """
    if not os.path.exists(caminho_excel):
        print(f"Arquivo não encontrado: {caminho_excel}")
        return

    # Abrir o Excel e ler os CNPJs da coluna 2
    workbook = openpyxl.load_workbook(caminho_excel)
    sheet = workbook.active

    # Lista para armazenar os CNPJs e e-mails encontrados
    resultados = []

    print(f"Iniciando busca de e-mails para CNPJs do site: {site_origem}...")
    
    for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):  # Começa na linha 2, coluna 2
        cnpj = row[1]
        if cnpj:
            print(f"Buscando e-mail para CNPJ: {cnpj}")
            email = buscar_email_cnpja(cnpj)
            resultados.append((cnpj, email))
    
    # Salvar os resultados em um novo Excel
    salvar_emails_excel(resultados, site_origem)
    print(f"Busca concluída! Resultados salvos.")

def buscar_email_cnpja(cnpj):
    """
    Faz a busca de um CNPJ no site 'cnpja.com/office/{CNPJ}' e retorna o e-mail encontrado.
    """
    url = f"https://cnpja.com/office/{cnpj}"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36"
    }

    try:
        response = requests.get(url, headers=headers, timeout=10)
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')

            # Extrai o e-mail da página (ajuste conforme a estrutura real do site)
            email_tag = soup.find("a", href=lambda href: href and "mailto:" in href)
            if email_tag:
                email = email_tag.get("href").replace("mailto:", "").strip()
                return email
            else:
                return "E-mail não encontrado"
        else:
            print(f"Erro ao acessar {url} (Status: {response.status_code})")
            return "Erro na requisição"
    except Exception as e:
        print(f"Erro ao buscar CNPJ {cnpj}: {e}")
        return "Erro ao buscar"

def salvar_emails_excel(resultados, site_origem):
    """
    Salva os resultados (CNPJs e e-mails) em um arquivo Excel.
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["CNPJ", "E-mail"])

    for cnpj, email in resultados:
        sheet.append([cnpj, email])

    nome_arquivo = f"emails_{site_origem}.xlsx"
    workbook.save(nome_arquivo)
    print(f"Resultados salvos em: {nome_arquivo}")

# Exemplo de integração com base no site selecionado
def executar_busca_emails(site_origem):
    """
    Determina qual arquivo Excel usar com base no site clicado.
    """
    arquivos_sites = {
        "site1": "dados_site1.xlsx",
        "site2": "dados_site2.xlsx",
        "site3": "dados_site3.xlsx",
        "site4": "dados_site4.xlsx",
        "portal_nacional": "dados_pncp.xlsx"
    }

    caminho_excel = arquivos_sites.get(site_origem)
    if caminho_excel:
        buscar_emails_por_cnpj(caminho_excel, site_origem)
    else:
        print("Site não reconhecido!")


buscar_emails_por_cnpj()