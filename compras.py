import openpyxl
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from tkinter import messagebox
from time import sleep
import random
import customtkinter as ctk
import re
from selenium.webdriver.common.action_chains import ActionChains

def verifica_ajudicacao(driver, root, label_status, cnpjs_set):
    situacao_texto = ""
    adjudicada_encontrada = False
    try:
        elementos_situacao = WebDriverWait(driver, 5).until(
            EC.visibility_of_all_elements_located((By.CSS_SELECTOR, "div[data-test='situacao-proposta']"))
        )

        for elemento_situacao in elementos_situacao:
            situacao_texto = elemento_situacao.text.strip()
            if situacao_texto.lower() == "adjudicada":  # Só pega o CNPJ se a situação for "adjudicada"
                adjudicada_encontrada = True
                print("A proposta está adjudicada. Realizando ações subsequentes...")

                try:
                    container = elemento_situacao.find_element(By.XPATH, "./ancestor::div[contains(@data-test, 'propostaItemEmSelecaoFornecedores')]")
                    container_cnpj = WebDriverWait(container, 10).until(
                        EC.visibility_of_element_located((By.XPATH, ".//span[@data-test='identificacao-participante']"))
                    )
                    cnpj = container_cnpj.text
                    print('CNPJ encontrado:', cnpj)

                    # Adiciona ao set para garantir que não haja duplicação
                    cnpjs_set.add(cnpj)

                    # Salva os CNPJs no Excel e atualiza o status
                    total_cnpjs = salvar_cnpjs_excel(cnpjs_set)
                    label_status.configure(text=f"Total de CNPJs encontrados: {total_cnpjs}")
                    print(f"Total de CNPJs encontrados: {total_cnpjs}")
                    root.update()

                except Exception as e:
                    print(f"Erro ao encontrar o botão de expansão: {e}")
                    continue
    except Exception as e:
        print(f"Erro ao buscar elementos de situação: {e}")
        return situacao_texto, False  # Caso não encontrar a situação

    return situacao_texto, adjudicada_encontrada

def raspar_pagina(driver, wb, root, label_status, cnpjs_set):
     # Coleta os links únicos
    elements = driver.find_elements(By.XPATH, "//i[@class='fa fa-tasks']")
    qtde_apps_card = len(driver.find_elements(By.XPATH, "//i[@class='fa fa-tasks']"))

    for index, element in enumerate(elements, start=0):
        

        elements = driver.find_elements(By.XPATH, "//i[@class='fa fa-tasks']")
        element = elements[index]
        #reencontra o elemento
        element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable(element)
        )

        url_inicial = driver.current_url

        print(f"Elemento {index+1} de {len(elements)}")
        driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", element)
        action = ActionChains(driver)
        action.move_to_element(element).click().perform()   
        print('clicado')  
        action.move_to_element(element).click().perform()  
        print('clicado dnv')
        print(qtde_apps_card)
        
        sleep(random.uniform(2, 3))

        url_atual = driver.current_url
        
        elements_details = driver.find_elements(By.XPATH, "//i[@class='fa-tasks fas']")

        if not elements_details and url_inicial != url_atual:
             print('Nao achei element details mas a URL mudou, voltando de pagina')
             driver.back()
             sleep(3)
             continue
        if elements_details and url_inicial != url_atual:
            print('Entrei em element details')
            for index_details, elementt in enumerate(elements_details, start=1):
                print(f"Clicando no elemento {index_details} de {len(elements_details)}")
                cont_element_details = len(elements_details)
                

                # Rolando suavemente até o elemento
                driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", elementt)
                

                # Ação para mover até o elemento e clicar
                action = ActionChains(driver)
                action.move_to_element(elementt).click().perform()

                print(f'Elemento {index} clicado com sucesso.')
                sleep(3)
                
                situacao_texto, adjudicada_encontrada = verifica_ajudicacao(driver, root, label_status, cnpjs_set)
                
                if not adjudicada_encontrada:
                        print('Nenhuma proposta adjudicada')
                        driver.back()
                        sleep(2)
                        driver.back()
                        sleep(3)
                else:
                    print('tem apenas 1 item, voltando pra pagina inicial')
                    driver.back()
                    sleep(2)
                    driver.back()  # Segunda vez
                    sleep(2)

                next_page = WebDriverWait(driver,5).until(
                    EC.element_to_be_clickable((By.XPATH,"//button[@class='p-ripple p-element p-paginator-next p-paginator-element p-link']"))
                )
                
                print(f'Elemento atual: {index}, último elemento: {qtde_apps_card - 1}')

                if index == qtde_apps_card - 1:
                    print(f'Elemento atual: {index}, último elemento: {(qtde_apps_card) - 1}')
                    next_page.click()
                    sleep(3)
                    raspar_pagina(driver, wb, root, label_status, cnpjs_set)
                else:
                    break
                    
                        
        else:
            print("Nenhum elemento encontrado.")
            continue

    wb.save("dados_vencedores_portal_compras_publicas.xlsx")
    sleep(15)
    driver.quit()
    messagebox.showinfo("Concluído", "Raspagem concluída com sucesso!")

def iniciar_raspagem_compras_gov(ano, root, label_status, cnpjs):
    # Inicializa o navegador com undetected_chromedriver
    driver = uc.Chrome()
    driver.get('https://cnetmobile.estaleiro.serpro.gov.br/comprasnet-web/public/compras')
    driver.maximize_window()
    sleep(random.uniform(1, 3))  # Delay aleatório

    # Criar a planilha para salvar os dados
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Link", "CNPJ", "Razão Social"])  

    # Interage com o filtro
    filtro = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, "//div[@class='p-radiobutton-box']"))
    )
    filtro.click()
    sleep(random.uniform(1, 3))  # Delay aleatório


    input_ano = WebDriverWait(driver,5).until(
        EC.element_to_be_clickable((By.XPATH,"//input[@placeholder='Ex: 102021']"))
    )
    input_ano.send_keys(ano)
    sleep(random.uniform(1, 3))

        # Clica no botão pesquisar
    button_pesquisar = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, "//button[@class='br-button is-primary']"))
    )
    button_pesquisar.click()
    sleep(random.uniform(1, 3))  # Delay aleatório

    raspar_pagina(driver, wb, root, label_status, cnpjs)

    wb.save("dados_vencedores_portal_compras_publicas.xlsx")
    sleep(5)
    driver.quit()
    messagebox.showinfo("Concluído", "Raspagem concluída com sucesso!")    


def salvar_cnpjs_excel(cnpjs_set):
    # Carregar o arquivo Excel existente ou criar um novo
    try:
        wb = openpyxl.load_workbook("dados_compras_gov.xlsx")
        ws = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CNPJs Extraídos"
        ws.append([None, "CNPJ"])  # Cabeçalho do arquivo Excel

    # Adiciona os CNPJs do set (sem duplicados)
    for cnpj in cnpjs_set:
        # Verifica se o CNPJ já está na planilha antes de adicionar
        cell_values = [cell.value for cell in ws['B']]  # Lista de valores existentes na coluna B
        if cnpj not in cell_values:
            ws.append([None, cnpj])  # Adiciona o CNPJ apenas se não estiver na planilha

    from datetime import datetime
    # Salva o arquivo
    data_atual = datetime.now().strftime("%d-%m-%Y")
    nome_arquivo = f"dados_compras_gov_{data_atual}.xlsx"
    wb.save(nome_arquivo)
    print(f"Total de CNPJs únicos salvos no arquivo Excel: {len(cnpjs_set)}")

    return len(cnpjs_set)  # Retorna a quantidade de CNPJs únicos adicionados

def criar_interface_compras_gov():
    ctk.set_appearance_mode("dark")
    ctk.set_default_color_theme("blue")

    root = ctk.CTk()
    root.geometry("400x300")
    root.title("Raspagem Compras Gov")

    ctk.CTkLabel(root, text="Digite o ano (Ex: 2023):").pack(pady=10)
    entry_ano = ctk.CTkEntry(root, placeholder_text="Digite o ano")
    entry_ano.pack(pady=10)

    label_status = ctk.CTkLabel(root, text="Status: Aguardando ação...", font=("Arial", 14))
    label_status.pack(pady=10)

    cnpjs_set = set()  # Inicializa o set de CNPJs aqui

    def iniciar():
        ano = entry_ano.get()
        if ano:
            iniciar_raspagem_compras_gov(ano, root, label_status, cnpjs_set)
        else:
            messagebox.showwarning("Erro", "Por favor, insira um ano.")

    ctk.CTkButton(root, text="Iniciar Raspagem", command=iniciar).pack(pady=10)
    root.mainloop()


criar_interface_compras_gov()