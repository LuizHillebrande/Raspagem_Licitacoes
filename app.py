import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import customtkinter as ctk
from tkinter import messagebox
import os

# Configurar o tema dark
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("dark-blue")

# Lista de sites de licitação/contratos
sites_licitacao = ["Portal Nacional de Contratações Públicas", "Outros Sites de Licitação"]


# Função para realizar a raspagem do Portal Nacional de Contratações Públicas
def iniciar_raspagem_Portal_Nacional_de_Contratacoes_Publicas():
    driver = webdriver.Chrome()

    # Criar a planilha para salvar os dados
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Link", "CNPJ", "Razão Social"])  # Cabeçalho da planilha

    for pagina in range(1, 1000):  # De 1 a 999
        url = f'https://pncp.gov.br/app/contratos?q=&pagina={pagina}'
        driver.get(url)
        wb.save("dados_fornecedores.xlsx")

        # Aguardar o carregamento dos elementos
        time.sleep(3)  # Espera de 3 segundos para garantir que a página carregue

        try:
            # Encontrar os botões únicos
            botoes = driver.find_elements(By.XPATH, "//a[contains(@class, 'br-item') and contains(@title, 'Acessar item.')]")
            botoes_unicos = list(dict.fromkeys([botao.get_attribute('href') for botao in botoes]))  # Remove duplicatas
            
            print(f"Página {pagina}: Total de botões encontrados: {len(botoes_unicos)}")

            for i, link in enumerate(botoes_unicos, start=1):
                try:
                    driver.get(link)
                    time.sleep(2)  # Espera um pouco para a página carregar

                    try:
                        # Extrair o CNPJ
                        cnpj_element = WebDriverWait(driver, 5).until(
                            EC.presence_of_element_located((By.XPATH, "//strong[contains(text(), 'CNPJ/CPF:')]/following-sibling::span"))
                        )
                        cnpj = cnpj_element.text
                        print(f"Botão {i}: CNPJ encontrado - {cnpj}")
                    except Exception as e:
                        print(f"Botão {i}: Erro ao extrair o CNPJ - {e}")
                        cnpj = "Não encontrado"

                    try:
                        # Extrair a Razão Social
                        razao_social_element = WebDriverWait(driver, 5).until(
                            EC.presence_of_element_located((By.XPATH, "//strong[contains(text(), 'Nome/Razão social:')]/following-sibling::span"))
                        )
                        razao_social = razao_social_element.text
                        print(f"Botão {i}: Razão social encontrada - {razao_social}")
                    except Exception as e:
                        print(f"Botão {i}: Erro ao extrair a Razão Social - {e}")
                        razao_social = "Não encontrado"

                    # Salvar os dados na planilha
                    ws.append([link, cnpj, razao_social])
                except Exception as e:
                    print(f"Erro ao processar o botão {i}: {e}")
          
        except Exception as e:
            print(f"Erro ao acessar a página {pagina}: {e}")

    # Salvar a planilha
    wb.save("dados_fornecedores.xlsx")
    driver.quit()
    messagebox.showinfo("Concluído", "Raspagem concluída com sucesso!")

def iniciar_raspagem_Portal_de_compras_publicas():
    driver = webdriver.Chrome()
    driver.get('https://www.portaldecompraspublicas.com.br/')

    # Criar a planilha para salvar os dados
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Link", "CNPJ", "Razão Social"])  

    time.sleep(3)  
    WebDriverWait(driver,5).until(
        EC.element_to_be_clickable((By.XPATH,"//a[@href class='busca-av']"))
    )
            
    wb.save("dados_vencedores_portal_compras_publicas.xlsx")
    driver.quit()
    messagebox.showinfo("Concluído", "Raspagem concluída com sucesso!")

#ELIMINAR CNPJ REPETIDO
def carregar_cnpjs_de_arquivos(arquivos):
    cnpjs_unicos = set()  # Usando um set para garantir que não haja CNPJs duplicados

    for arquivo in arquivos:
        wb = openpyxl.load_workbook(arquivo)
        ws = wb.active
        
       
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):  # Pular o cabeçalho
            cnpj = row[1] 
            if cnpj:
                cnpjs_unicos.add(cnpj.strip())  # Adiciona o CNPJ ao set (remover espaços extras)

    return cnpjs_unicos

def salvar_cnpjs_unicos(cnpjs_unicos, nome_arquivo="cnpjs_unicos.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["CNPJ"])  # Cabeçalho da planilha

    for cnpj in cnpjs_unicos:
        ws.append([cnpj])

    wb.save(nome_arquivo)
    print(f"CNPJs únicos salvos em: {nome_arquivo}")

# Configurar o tema dark
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("dark-blue")

# Funções de cada ação (simplificadas para este exemplo)
def iniciar_raspagem():
    open_site_selection_window()

def eliminar_cnpjs():
     # Lista de arquivos que devem ser processados
    arquivos = ["dados_fornecedores.xlsx", "dados_fornecedores_2.xlsx"]

    # Verifica se os arquivos existem
    arquivos_existentes = [arquivo for arquivo in arquivos if os.path.exists(arquivo)]

    if not arquivos_existentes:
        # Exibe mensagem de aviso e retorna
        messagebox.showwarning("Atenção", "Nenhum arquivo de dados foi encontrado para eliminar CNPJs duplicados.")
        return

    cnpjs_unicos = carregar_cnpjs_de_arquivos(arquivos_existentes)


    salvar_cnpjs_unicos(cnpjs_unicos)
    messagebox.showinfo("Concluído", "CNPJs duplicados eliminados com sucesso!")

def enviar_emails():
    messagebox.showinfo("Enviar E-mails", "Iniciando envio de e-mails.")

def raspar_emails():
    # Função que abre a janela de seleção para raspagem de e-mails
    open_email_selection_window()

def open_site_selection_window():
    # Criar a nova janela para selecionar o site
    site_window = ctk.CTkToplevel()
    site_window.title("Selecionar Site para Raspar Dados")
    site_window.geometry("400x300")
    
    # Adicionar as opções de sites
    ctk.CTkLabel(site_window, text="Selecione o Site", font=("Arial", 16)).pack(pady=20)
    
    # Lista de sites para selecionar
    sites_licitacao = ["Portal Nacional de Contratações Públicas", "Outros Sites de Licitação"]
    
    # Combobox para selecionar o site
    combo_sites = ctk.CTkComboBox(site_window, values=sites_licitacao, width=200)
    combo_sites.pack(pady=10)
    
    
    def confirm_site_selection():
        site = combo_sites.get()
        if site:
            if site == 'Portal Nacional de Contratações Públicas':
                iniciar_raspagem_Portal_Nacional_de_Contratacoes_Publicas()
        else:
            messagebox.showwarning("Atenção", "Por favor, selecione um site.")

    btn_confirmar = ctk.CTkButton(site_window, text="Confirmar Seleção", command=confirm_site_selection)
    btn_confirmar.pack(pady=10)

def open_email_selection_window():
    # Criar a nova janela para selecionar a opção de raspagem de e-mails
    email_window = ctk.CTkToplevel()
    email_window.title("Selecionar Opção para Raspagem de E-mails")
    email_window.geometry("400x300")
    
    # Adicionar as opções de e-mail
    ctk.CTkLabel(email_window, text="Selecione a Ação de E-mails", font=("Arial", 16)).pack(pady=20)
    
    # Combobox para selecionar a ação de e-mails
    email_options = ["Raspar E-mails de Sites", "Importar Lista de E-mails"]
    
    combo_emails = ctk.CTkComboBox(email_window, values=email_options, width=200)
    combo_emails.pack(pady=10)
    
    # Botão para confirmar a seleção
    def confirm_email_selection():
        email_action = combo_emails.get()
        if email_action:
            messagebox.showinfo("Ação Selecionada", f"Você selecionou: {email_action}")
            email_window.destroy()  # Fechar a janela de seleção de e-mail
        else:
            messagebox.showwarning("Atenção", "Por favor, selecione uma ação de e-mail.")

    btn_confirmar_email = ctk.CTkButton(email_window, text="Confirmar Seleção", command=confirm_email_selection)
    btn_confirmar_email.pack(pady=10)

# Criar a janela principal
root = ctk.CTk()
root.title("Sistema de Licitações")
root.geometry("600x400")
root.resizable(False, False)

# Centralizar componentes
root.grid_rowconfigure(0, weight=1)
root.grid_rowconfigure(1, weight=1)
root.grid_rowconfigure(2, weight=1)
root.grid_rowconfigure(3, weight=1)
root.grid_columnconfigure(0, weight=1)

# Título
ctk.CTkLabel(root, text="Selecione o que deseja fazer", font=("Arial", 18)).grid(row=0, column=0, pady=20)

# Tabela com opções
opcoes = [
    ("Raspar Dados", iniciar_raspagem),
    ("Eliminar CNPJ Repetido", eliminar_cnpjs),
    ("Enviar E-mails", enviar_emails),
    ("Raspar E-mails", raspar_emails),  
]

for idx, (texto, comando) in enumerate(opcoes):
    ctk.CTkButton(root, text=texto, command=comando, width=200).grid(row=1+idx, column=0, pady=10)

# Rodapé
ctk.CTkLabel(root, text="Desenvolvido por Luiz Fernando Hillebrande - Todos os direitos reservados", font=("Arial", 10), anchor="center").grid(row=4, column=0, pady=10)

# Rodar o loop principal
root.mainloop()
