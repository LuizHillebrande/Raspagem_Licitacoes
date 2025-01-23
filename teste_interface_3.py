import customtkinter as ctk
from tkinter import messagebox
from PIL import Image
import openpyxl
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import time
import pyautogui
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import StaleElementReferenceException, ElementClickInterceptedException
from selenium.webdriver.common.action_chains import ActionChains
import os
import re
import pdfplumber
import pandas as pd
import customtkinter as ctk
import json
import threading
from selenium.common.exceptions import StaleElementReferenceException
from selenium.common.exceptions import NoSuchElementException
from time import sleep
import tkinter as tk
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import customtkinter as ctk
from tkinter import messagebox
import os
import threading
from datetime import datetime
import logging

import customtkinter as ctk
from tkinter import messagebox
from PIL import Image
import openpyxl
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import time
import pyautogui
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import StaleElementReferenceException, ElementClickInterceptedException
from selenium.webdriver.common.action_chains import ActionChains
import os
import re
import pdfplumber
import pandas as pd
import customtkinter as ctk
import json
import threading
from selenium.common.exceptions import StaleElementReferenceException
from selenium.common.exceptions import NoSuchElementException
from time import sleep
import tkinter as tk
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
from openpyxl import load_workbook
from tkinter import filedialog
import cv2
import numpy as np

caminho_arquivo_selecionado = ""

# Declaração global de entry_arquivo
entry_arquivo = None

def converter_para_csv(arquivo_excel):
    try:
        # Carrega o arquivo Excel
        df = pd.read_excel(arquivo_excel)

        # Cria a pasta "emails" no diretório atual, se não existir
        pasta_destino = "./emails"
        if not os.path.exists(pasta_destino):
            os.makedirs(pasta_destino)

        # Define o nome do arquivo CSV dentro da pasta "emails"
        nome_arquivo = os.path.basename(arquivo_excel)  # Nome do arquivo Excel
        nome_csv = os.path.splitext(nome_arquivo)[0] + ".csv"
        caminho_csv = os.path.join(pasta_destino, nome_csv)

        # Salva o arquivo como CSV na pasta "emails"
        df.to_csv(caminho_csv, index=False, encoding='utf-8')
        messagebox.showinfo("Sucesso", f"Arquivo convertido para CSV com sucesso!\nSalvo em: {caminho_csv}")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao converter o arquivo: {e}")

def selecionar_e_converter():
    # Abre o explorador de arquivos para selecionar o arquivo Excel
    arquivo_selecionado = filedialog.askopenfilename(
        title="Selecione um arquivo Excel",
        filetypes=(("Arquivos Excel", "*.xlsx *.xls"), ("Todos os Arquivos", "*.*"))
    )
    
    if arquivo_selecionado:
        # Chama a função para converter o arquivo selecionado
        converter_para_csv(arquivo_selecionado)

def enviar_emails():
    driver = webdriver.Chrome()
    driver.get('https://marilirequiaseguros.com.br/mautic/s/login')
    driver.maximize_window()


    login = 'luiz.logika@gmail.com'
    password = 'Dev123@'
    logar = WebDriverWait(driver,5).until(
        EC.element_to_be_clickable((By.XPATH,"//input[@id='username']"))
    )
    logar.send_keys(login)

    password_enter = WebDriverWait(driver,5).until(
        EC.element_to_be_clickable((By.XPATH,"//input[@type='password']"))
    )
    password_enter.send_keys(password)

    enter = WebDriverWait(driver,5).until(
        EC.element_to_be_clickable((By.XPATH,"//button[@class='btn btn-lg btn-primary btn-block']"))
    )
    enter.click()

    sleep(4)


    contacts = WebDriverWait(driver,5).until(
        EC.element_to_be_clickable((By.XPATH,"//a[@data-menu-link='mautic_contact_index']"))
    )
    contacts.click()
    sleep(3)

    seta = WebDriverWait(driver,5).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR,"i[class='fa fa-caret-down']"))
    )
    seta.click()
    sleep(2)

    importar = WebDriverWait(driver,5).until(
        EC.element_to_be_clickable((By.XPATH,"//a[@href='/mautic/s/contacts/import/new']"))
    )
    importar.click()
    sleep(2)


    limite = WebDriverWait(driver,5).until(
        EC.element_to_be_clickable((By.XPATH,"//input[@id='lead_import_batchlimit']"))
    )
    limite.clear()
    limite.send_keys('2000')
    sleep(2)

    import_csv = WebDriverWait(driver,5).until(
        EC.element_to_be_clickable((By.XPATH,"//input[@id='lead_import_file']"))
    )
    import_csv.send_keys(caminho_arquivo_selecionado)
    sleep(2)

    carregar = WebDriverWait(driver,5).until(
        EC.element_to_be_clickable((By.XPATH,"//button[@id='lead_import_start']"))
    )
    carregar.click()

    propriedade_contato = WebDriverWait(driver, 10).until(
    EC.presence_of_element_located((By.XPATH, "//div[@id='lead_field_import_owner_chosen']"))
    )
    propriedade_contato.click()
    sleep(1)

    opcao = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, "//li[contains(text(), 'Scraping, Teste')]"))
    )
    opcao.click()

    segmento_contato = WebDriverWait(driver, 10).until(
    EC.presence_of_element_located((By.XPATH, "//div[@id='lead_field_import_list_chosen']"))
    )
    segmento_contato.click()
    sleep(1)

    
    option_segmento = WebDriverWait(driver, 5).until(
        EC.presence_of_element_located((By.XPATH, "//li[contains(text(), 'teste_mautic')]"))  # Ajuste o texto conforme necessário
    )
    option_segmento.click()

    # Selecionar a coluna de email
    email_coluna = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "//div[@id='lead_field_import_email_chosen']"))
    )
    email_coluna.click()
    sleep(1)

    # Selecionar a opção de email
    option_email = WebDriverWait(driver, 5).until(
        EC.presence_of_element_located((By.XPATH, "//li[contains(text(), 'Email')]"))  
    )
    option_email.click()
    sleep(3)

    # Selecionar a coluna de razão social
    #razao_social_coluna = WebDriverWait(driver, 5).until(
        #EC.presence_of_element_located((By.XPATH, "//div[@id='lead_field_import_razao_social_chosen']"))
    #)
    #razao_social_coluna.click()

    pyautogui.click(992,616, duration=1)
    sleep(1)

    pyautogui.write('Company Name')
    sleep(1)

    # Enviar "Enter" para confirmar
    pyautogui.press('enter')
    sleep(2)

    #importacao_plano_fundo  = WebDriverWait(driver,5).until(
        #EC.element_to_be_clickable((By.XPATH,"//button[@id='lead_field_import_buttons_apply_toolbar']"))
    #)
    #importacao_plano_fundo.click()

    importacao_navegador = WebDriverWait(driver,5).until(
        EC.element_to_be_clickable((By.XPATH,"//button[@id='lead_field_import_buttons_save_toolbar']"))
    )
    importacao_navegador.click()
    sleep(2)

    nome_arquivo = os.path.basename(caminho_arquivo_selecionado)

    #link_arquivo = WebDriverWait(driver, 10).until(
        #EC.presence_of_element_located((By.XPATH, f"//a[contains(@href, 'view/') and contains(text(), '{nome_arquivo}')]"))
    #)

    # Agora, monitoramos o progresso
    monitorar_elemento(xpath, url, driver)

    
    sleep(15)
    driver.quit()

def monitorar_elemento(xpath, url, driver):
    while True:
        try:
            # Procura pelo elemento usando o XPath
            elemento = driver.find_element(By.XPATH, xpath)
            if elemento.is_displayed():
                print("Elemento encontrado! Texto:", elemento.text)
                sleep(5)
                driver.get('https://marilirequiaseguros.com.br/mautic/s/login')
                driver.maximize_window()

                sleep(2)
                canais_dropdown = WebDriverWait(driver,5).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, "a#mautic_channels_root .arrow.pull-right.text-right"))
                )
                canais_dropdown.click()
                sleep(2)

                emails = WebDriverWait(driver,5).until(
                    EC.element_to_be_clickable((By.XPATH,"//a[@data-menu-link='mautic_email_index']"))
                )
                emails.click()
                sleep(5)

                tr_element = driver.find_element(By.XPATH, "//td[text()='38']/ancestor::tr")
                button = tr_element.find_element(By.CSS_SELECTOR, "button.dropdown-toggle")  # Encontra o botão dentro do tr
                WebDriverWait(driver, 10).until(EC.element_to_be_clickable(button)).click()
                sleep(2)

                #link_enviar = driver.find_element(By.XPATH, "//td[text()='38']/ancestor::tr//a[@href='/mautic/s/emails/send/38']")

                # Garantir que o link seja clicável e clicar nele
                #WebDriverWait(driver, 10).until(EC.element_to_be_clickable(link_enviar)).click()
                pyautogui.click(337,497,duration=1)
                sleep(3)
                pyautogui.click(1221,402,duration=1)
                sleep(3)
                pyautogui.click(988,305,duration=1)
                sleep(3)
                

                monitorar_elemento_envio_emails(xpath, url, driver)
                break
        except NoSuchElementException:
            print("Elemento não encontrado, tentando novamente em 30 segundos...")

        # Aguarda 30 segundos antes de tentar novamente
        time.sleep(10)

def monitorar_elemento_envio_emails(xpath, url, driver):
    while True:
        try:
            # Procura pelo elemento usando o XPath
            elemento = driver.find_element(By.XPATH, xpath)
            if elemento.is_displayed():
                print("Elemento encontrado! Texto:", elemento.text)
                messagebox.showinfo('Sucesso!', 'E-mails enviados!')
                driver.quit
                break
        except NoSuchElementException:
            print("Elemento não encontrado, tentando novamente em 30 segundos...")

        # Aguarda 30 segundos antes de tentar novamente
        time.sleep(30)

# Configurações
url = "URL_DA_SUA_PAGINA"  # Substitua pela URL da página a ser monitorada
xpath = '//div[@class="panel-heading"]//h4[contains(text(), "Sucesso!")]'

def abrir_pasta_emails():
    global entry_arquivo
    global caminho_arquivo_selecionado
    # Abre o explorador de arquivos na pasta "emails"
    pasta_emails = './emails'
    if not os.path.exists(pasta_emails):
        print("A pasta 'emails' não existe.")
        return
    arquivo_selecionado = filedialog.askopenfilename(
    initialdir=pasta_emails,
    title="Escolha um arquivo CSV",
    filetypes=(("CSV Files", "*.csv"),)
)
    
    if arquivo_selecionado:
        global caminho_arquivo_selecionado
        caminho_arquivo_selecionado = arquivo_selecionado
        entry_arquivo.delete(0, ctk.END)  # Limpa a entrada anterior
        entry_arquivo.insert(0, arquivo_selecionado)  # Insere o caminho do arquivo selecionado

def fazer_upload():
    if caminho_arquivo_selecionado:
        print(f"Arquivo selecionado para upload: {caminho_arquivo_selecionado}")
        enviar_emails()  # Chama a função de envio dos emails
    else:
        print("Nenhum arquivo selecionado!")

def criar_interface_mautic():
    global entry_arquivo
    root = ctk.CTk()
    root.title("Upload de Arquivo Excel")
    root.geometry("600x400")

    label_titulo = ctk.CTkLabel(root, text="Escolha o arquivo Excel para upload", font=("Arial", 14))
    label_titulo.pack(pady=10)

    entry_arquivo = ctk.CTkEntry(root, width=300)
    entry_arquivo.pack(pady=5)

    botao_selecionar = ctk.CTkButton(
        root, 
        text="Converter para CSV", 
        command=selecionar_e_converter,
        font=("Arial", 14)
    )
    botao_selecionar.pack(pady=60)

    botao_abrir_pasta = ctk.CTkButton(root, text="Abrir Pasta 'emails'", command=abrir_pasta_emails)
    botao_abrir_pasta.pack(pady=5)

    botao_upload = ctk.CTkButton(root, text="Fazer Upload", command=fazer_upload)
    botao_upload.pack(pady=10)

    root.mainloop()


logging.basicConfig(level=logging.ERROR, format='%(asctime)s - %(levelname)s - %(message)s')


data_mais_recente = None
data_mais_antiga = None
intervalo_label = None

def atualizar_intervalo_datas(datas):
    global data_mais_recente, data_mais_antiga
    if datas:
        data_atual_mais_recente = max(datas)
        data_atual_mais_antiga = min(datas)
        
        # Atualizar as variáveis globais
        if not data_mais_recente or data_atual_mais_recente > data_mais_recente:
            data_mais_recente = data_atual_mais_recente
        if not data_mais_antiga or data_atual_mais_antiga < data_mais_antiga:
            data_mais_antiga = data_atual_mais_antiga

        # Atualizar o Label na interface gráfica
        intervalo_label.configure(
            text=f"Filtrado de {data_mais_antiga.strftime('%d/%m/%Y')} até {data_mais_recente.strftime('%d/%m/%Y')}"
        )

def extrair_datas_da_pagina(driver):
    datas_extradas = []
    try:
        elementos = driver.find_elements(By.XPATH, "//div[@class='datatable-body-cell-label']")
        for elemento in elementos:
            texto = elemento.text.strip()
            try:
                # Tentar converter o texto em uma data
                data = datetime.strptime(texto, "%d/%m/%Y")
                datas_extradas.append(data)
            except ValueError:
                # Ignorar valores que não sejam datas
                continue
    except Exception as e:
        logging.error(f"Erro ao extrair datas: {e}")
    return datas_extradas

def extrair_data_assinatura(driver):
    try:
        strong_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//strong[text()='Data de assinatura:']"))
        )

        # Encontrar o <span> imediatamente após o <strong> encontrado
        span_data_element = strong_element.find_element(By.XPATH, "following-sibling::span")

        # Extrair o texto da data
        data_texto = span_data_element.text.strip()

        # Retornar a data extraída
        return data_texto
    except Exception as e:
        logging.error(f"Erro ao extrair a data de assinatura: {e}")
        return "Não encontrada"

# Configurar o tema dark
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("dark-blue")

cnpj_count = 0

# Lista de sites de licitação/contratos
sites_licitacao = ["Portal Nacional de Contratações Públicas", "Outros Sites de Licitação"]


# Função para realizar a raspagem do Portal Nacional de Contratações Públicas
def iniciar_raspagem_Portal_Nacional_de_Contratacoes_Publicas():
    global cnpj_count
    driver = webdriver.Chrome()

    data_atual = datetime.now().strftime("%d-%m-%Y")
    nome_arquivo = f"dados_fornecedores_{data_atual}.xlsx"  # Nome com data no final

    # Criar a planilha para salvar os dados
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Link", "CNPJ", "Razão Social", "Data de Divulgação no PNCP", "Data de Assinatura"])

    for pagina in range(1, 1000):  # De 1 a 999
        url = f'https://pncp.gov.br/app/contratos?q=&pagina={pagina}'
        driver.get(url)

        # Aguardar o carregamento dos elementos
        time.sleep(3)  # Espera de 3 segundos para garantir que a página carregue

        try:
            # Encontrar os botões únicos
            botoes = driver.find_elements(By.XPATH, "//a[contains(@class, 'br-item') and contains(@title, 'Acessar item.')]")
            botoes_unicos = list(dict.fromkeys([botao.get_attribute('href') for botao in botoes]))  # Remove duplicatas
            
            print(f"Página {pagina}: Total de botões encontrados: {len(botoes_unicos)}")

            for i, link in enumerate(botoes_unicos, start=1):
                try:
                    driver.get(link)
                    time.sleep(2)  # Espera um pouco para a página carregar

                    try:
                        # Extrair o CNPJ
                        cnpj_element = WebDriverWait(driver, 5).until(
                            EC.presence_of_element_located((By.XPATH, "//strong[contains(text(), 'CNPJ/CPF:')]/following-sibling::span"))
                        )
                        cnpj = cnpj_element.text
                        print(f"Botão {i}: CNPJ encontrado - {cnpj}")
                        cnpj_count += 1
                        cnpj_label.configure(text=f"CNPJs extraídos: {cnpj_count}")
                    except Exception as e:
                        print(f"Botão {i}: Erro ao extrair o CNPJ - {e}")
                        cnpj = "Não encontrado"

                    try:
                        # Extrair a Razão Social
                        razao_social_element = WebDriverWait(driver, 5).until(
                            EC.presence_of_element_located((By.XPATH, "//strong[contains(text(), 'Nome/Razão social:')]/following-sibling::span"))
                        )
                        razao_social = razao_social_element.text
                        print(f"Botão {i}: Razão social encontrada - {razao_social}")
                    except Exception as e:
                        print(f"Botão {i}: Erro ao extrair a Razão Social - {e}")
                        razao_social = "Não encontrado"

                    datas_pagina = extrair_datas_da_pagina(driver)
                    if datas_pagina:
                        data_divulgacao = datas_pagina[0].strftime("%d/%m/%Y")  # Usando a primeira data extraída
                    else:
                        data_divulgacao = "Não encontrada"
                    
                    data_texto = extrair_data_assinatura(driver)

                    # Atualizar intervalo de datas
                    atualizar_intervalo_datas(datas_pagina)

                    print(f"Data mais recente: {data_mais_recente.strftime('%d/%m/%Y')}")
                    print(f"Data mais antiga: {data_mais_antiga.strftime('%d/%m/%Y')}")
                    
                    
                    # Salvar os dados na planilha
                    ws.append([link, cnpj, razao_social, data_divulgacao, data_texto])
                    wb.save(nome_arquivo)

                except Exception as e:
                    print(f"Erro ao processar o botão {i}: {e}")
          
        except Exception as e:
            print(f"Erro ao acessar a página {pagina}: {e}")

    # Salvar a planilha
    
    driver.quit()
    messagebox.showinfo("Concluído", "Raspagem concluída com sucesso!")

def iniciar_raspagem():
    threading.Thread(target=iniciar_raspagem_Portal_Nacional_de_Contratacoes_Publicas, daemon=True).start()

# Criando a interface gráfica com CustomTkinter
def criar_interface_pncp():
    global cnpj_label
    global intervalo_label

    # Configuração da janela
    janela = ctk.CTk()
    janela.title("Raspagem de Dados de Licitação")
    janela.geometry("500x300")

    # Título
    titulo = ctk.CTkLabel(janela, text="Raspagem PNCP", font=("Arial", 18))
    titulo.pack(pady=20)

    # Botão para iniciar a raspagem
    btn_iniciar = ctk.CTkButton(janela, text="Iniciar Raspagem", command=iniciar_raspagem, font=("Arial", 16))
    btn_iniciar.pack(pady=20)

    # Label para mostrar o número de CNPJs extraídos
    cnpj_label = ctk.CTkLabel(janela, text="CNPJs extraídos: 0", font=("Arial", 14))
    cnpj_label.pack(pady=10)

    intervalo_label = ctk.CTkLabel(janela, text="Filtrado de --/--/---- até --/--/----", font=("Arial", 14))
    intervalo_label.pack(pady=10)

    # Iniciar a interface
    janela.mainloop()

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import time
import pyautogui
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import StaleElementReferenceException, ElementClickInterceptedException
from selenium.webdriver.common.action_chains import ActionChains
import os
import re
import pdfplumber
import pandas as pd
import customtkinter as ctk
import json

PROGRESSO_JSON = "progresso.json"

def salvar_progresso(pdfs_baixados, cnpjs_extraidos):
    progresso = {
        "pdfs_baixados": pdfs_baixados,
        "cnpjs_extraidos": cnpjs_extraidos
    }
    with open(PROGRESSO_JSON, "w") as file:
        json.dump(progresso, file)

# Função para carregar progresso
def carregar_progresso():
    if os.path.exists(PROGRESSO_JSON):
        with open(PROGRESSO_JSON, "r") as file:
            return json.load(file)
    return {"pdfs_baixados": 0, "cnpjs_extraidos": 0}

def limpa_campo():
    for _ in range(10):
        pyautogui.press('right')
    for _ in range(10):
        pyautogui.press('backspace')

def scroll_ate_resultados_estabilizarem(driver):
    # Obter o número inicial de resultados
    resultado_atual = driver.find_element(By.ID, "footResultCount").text
    resultado_atual = int(resultado_atual.split(":")[-1].strip())

    while True:
        time.sleep(1)
        # Rolar para baixo
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(4)  # Aguarda o carregamento dos novos resultados

        # Obter o novo número de resultados
        novo_resultado = driver.find_element(By.ID, "footResultCount").text
        novo_resultado = int(novo_resultado.split(":")[-1].strip())

        # Verificar se o número de resultados mudou
        if novo_resultado == resultado_atual:
            print(f"Número de resultados estabilizado: {novo_resultado}")
            break
        else:
            resultado_atual = novo_resultado  # Atualiza o número de resultados

def extrair_bllcompras(data_inicio, data_fim, status_processo,label_contador_pdfs):

    progresso_json = "progresso.json"

    if os.path.exists(progresso_json):
        os.remove(progresso_json)
        print(f"Arquivo {progresso_json} excluído.")

    current_dir = os.getcwd()
    if status_processo == "HOMOLOGADO":
        pasta_destino = os.path.join(current_dir, "vencedores_bll_compras_homologado")
    elif status_processo == "ADJUDICADO":
        pasta_destino = os.path.join(current_dir, "vencedores_bll_compras_adjudicado")
    else:
        print(f"Erro: Status de processo inválido: {status_processo}")
        return

    # Verifica se a pasta existe, se não, cria a pasta
    if not os.path.exists(pasta_destino):
        os.makedirs(pasta_destino)
        print(f"Pasta criada: {pasta_destino}")
    chrome_options = webdriver.ChromeOptions()
    prefs = {
        "download.default_directory": pasta_destino,  # Define o diretório atual
        "download.prompt_for_download": False,      # Não perguntar antes de baixar
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True                # Ativa o download seguro
    }
    chrome_options.add_experimental_option("prefs", prefs)

    driver = webdriver.Chrome(options=chrome_options)
    driver.get('https://bllcompras.com/Process/ProcessSearchPublic?param1=0#')
    driver.maximize_window()

    

    # Aguarda até que o elemento de seleção de status esteja presente
    select_element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "fkStatus"))
    )
    select = Select(select_element)
    select.select_by_visible_text(status_processo)

    # Aguarda até que o botão de busca avançada seja clicável
    busca_avancada = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.XPATH, "//a[@title='BUSCA AVANÇADA']"))
    )
    busca_avancada.click()

    public_inicio = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.XPATH, "//input[@data-target='#DateStart']"))
    )
    public_inicio.click()
    public_inicio.clear()
    public_inicio.send_keys(data_inicio)

    # Preencher a data final
    public_final = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.XPATH, "//input[@data-target='#DateEnd']"))
    )
    public_final.click()
    public_final.clear()
    public_final.send_keys(data_fim)

    # Aguarda até que o ícone de pesquisa seja clicável
    icon = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, "button i.fas.fa-search"))
    )
    icon.click()

    scroll_ate_resultados_estabilizarem(driver)

    # Espera até que a lista de elementos de informações esteja carregada
    WebDriverWait(driver, 10).until(
        EC.presence_of_all_elements_located((By.CSS_SELECTOR, "i.fas.fa-info-circle"))
    )

    original_window = driver.current_window_handle
    original_url = driver.current_url

    elements = driver.find_elements(By.CSS_SELECTOR, "i.fas.fa-info-circle")

    pdfs_baixados = 0
    
    progresso = carregar_progresso()
    pdfs_baixados = progresso["pdfs_baixados"]

    for index, element in enumerate(elements, start=1):
        print(f"Elemento {index} de {len(elements)}")

        elements = driver.find_elements(By.CSS_SELECTOR, "i.fas.fa-info-circle")
        element = elements[index]
        
        # Reencontrando o elemento antes de clicar para garantir que ele ainda é válido
        element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable(element)
        )
        driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", element)
        action = ActionChains(driver)
        action.move_to_element(element).click().perform()
        
        time.sleep(2)
        print("BOTAO 1 CLICADO")
        

        WebDriverWait(driver, 10).until(
        lambda driver: len(driver.window_handles) > 1  # Garantir que uma nova janela foi aberta
        )
        
        # Trocar para a nova janela
        new_window = [window for window in driver.window_handles if window != original_window][0]
        driver.switch_to.window(new_window)

        current_url = driver.current_url
        print("URL da nova janela:", current_url)

        try:
            # Exemplo: esperando por um botão que deve aparecer na nova janela
            relatorios_button = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, "//button[@title='Relatórios']"))
            )
            print("Elemento 'Relatórios' apareceu, a página foi carregada corretamente.")
            
            # Clicar no botão 'Relatórios' caso apareça
            relatorios_button.click()
            time.sleep(2)
            
        except Exception as e:
            print("O elemento 'Relatórios' não apareceu ou houve um erro:", e)
        
        try:
            # Aguarda a presença do elemento <b> e verifica o texto
            mensagem_elemento = WebDriverWait(driver, 3).until(
                EC.presence_of_element_located((By.XPATH, "//b"))
            )
            mensagem_texto = mensagem_elemento.text

            # Verifica o conteúdo do texto
            if "relatórios ainda não estão disponíveis" in mensagem_texto:
                print("Relatórios não estão disponíveis. Pulando para o próximo item...")
                pyautogui.hotkey('ctrl','w')
                driver.switch_to.window(original_window)
                continue  # Passa para o próximo elemento da lista
        except Exception:
            print("Elemento <b> não encontrado. Continuando o download...")

        WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.TAG_NAME, "tr"))
        )

        linhas = driver.find_elements(By.TAG_NAME, "tr")
    

        for linha in linhas:
            try:
                # Verifica se a linha contém o texto "VENCEDORES DO PROCESSO"
                if "VENCEDORES DO PROCESSO" in linha.text.upper():
                    print("Linha encontrada com 'VENCEDORES DO PROCESSO'.")

                    # Procura todas as tags <a> com o atributo 'download' dentro da linha
                    links_com_download = linha.find_elements(By.CSS_SELECTOR, "a[download]")
                    print(f"Total de links com 'download' encontrados: {len(links_com_download)}")

                    # Itera sobre os links com atributo 'download'
                    for link in links_com_download:
                        # Verifica se o atributo 'download' ou texto contém "VENCEDORES DO PROCESSO"
                        if "VencedoresProcesso" in link.get_attribute("download"):
                            print("Link correto encontrado com 'VencedoresProcesso' no atributo 'download'.")

                            try:
                                botao_download = link.find_element(By.CSS_SELECTOR, "i.fas.fa-download")
                                print("Botão de download encontrado. Tentando clicar...")

                                # Realiza o scroll até o botão para garantir que ele esteja visível
                                driver.execute_script("arguments[0].scrollIntoView(true);", botao_download)
                                time.sleep(1)

                                # Realiza o clique com ActionChains
                                ActionChains(driver).move_to_element(botao_download).click().perform()
                                print("Download iniciado com sucesso.")
                                pdfs_baixados += 1
                                label_contador_pdfs.configure(text=f"PDFs    Baixados: {pdfs_baixados}")
                                salvar_progresso(pdfs_baixados, progresso["cnpjs_extraidos"])
                                root.update()
                                time.sleep(5)  
                                download_realizado = True
                                break  

                            except Exception as e:
                                print(f"Erro ao tentar clicar no ícone de download: {e}")
                        if download_realizado:
                            print("Download concluído para esta linha. Encerrando busca na linha.")
                            break  
                        else:
                            print("Link ignorado, não corresponde ao 'VencedoresProcesso'.")

            except Exception as e:
                print(f"Erro ao processar a linha: {e}")

        # Fecha a aba atual e retorna à aba original
        pyautogui.hotkey('ctrl', 'w')
        driver.switch_to.window(original_window)

    # Espera um pouco mais para garantir que o download seja iniciado
    time.sleep(5)  # Pausa de 5 segundos antes de fechar o driver
    scroll_ate_resultados_estabilizarem(driver)
    driver.quit()  # Fechamento do driver após a execução

def extrair_cnpjs_pasta(pasta, nome_arquivo_saida, label_erro):
    """
    Extrai CNPJs de todos os PDFs em uma pasta específica e salva em um arquivo Excel.
    
    :param pasta: Caminho da pasta contendo os PDFs
    :param nome_arquivo_saida: Nome do arquivo Excel de saída onde os CNPJs serão salvos
    :param label_erro: Widget de erro para exibir mensagens na interface
    """
    # Padrão regex para capturar CNPJs
    padrao_cnpj = r'\b\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}\b'
    progresso = carregar_progresso()
    cnpjs_encontrados = set()
    total_cnpjs = progresso["cnpjs_extraidos"]

    # Itera sobre os arquivos na pasta
    for arquivo in os.listdir(pasta):
        if arquivo.endswith(".pdf"):
            caminho_arquivo = os.path.join(pasta, arquivo)
            print(f"Processando: {caminho_arquivo}")

            # Abre o PDF e extrai texto
            try:
                with pdfplumber.open(caminho_arquivo) as pdf:
                    for pagina in pdf.pages:
                        texto = pagina.extract_text()
                        if texto:
                            # Procura por CNPJs no texto
                            cnpjs = re.findall(padrao_cnpj, texto)
                            cnpjs_encontrados.update(cnpjs)
                            total_cnpjs += len(cnpjs)
                            label_contador_cnpjs.configure(text=f"CNPJs Extraídos: {total_cnpjs}")
                            salvar_progresso(progresso["pdfs_baixados"], total_cnpjs)
                            root.update()
                            if cnpjs:
                                print(f"CNPJs encontrados no arquivo {arquivo}: {cnpjs}")
                                cnpjs_encontrados.update(cnpjs)
            except Exception as e:
                print(f"Erro ao processar {arquivo}: {e}")

    # Remove duplicatas de CNPJs
    cnpjs_unicos = list(set(cnpjs_encontrados))

    # Exibe todos os CNPJs encontrados
    if cnpjs_unicos:
        print("\nCNPJs encontrados:")
        for cnpj in cnpjs_unicos:
            print(cnpj)
        
        # Salva os CNPJs em um arquivo Excel
        caminho_saida = os.path.join(os.getcwd(), nome_arquivo_saida)  # Caminho completo para o arquivo
        df = pd.DataFrame({
            "Índice": range(1, len(cnpjs_unicos) + 1),
            "CNPJs": cnpjs_unicos  # Insira os CNPJs na segunda coluna
        })
        df.to_excel(caminho_saida, index=False)
        print(f"\nCNPJs salvos em: {caminho_saida}")
        
        # Atualiza a interface com uma mensagem de sucesso
        label_erro.configure(text=f"CNPJs extraídos e salvos em {nome_arquivo_saida}", text_color="green")
    else:
        label_erro.configure(text="Nenhum CNPJ encontrado nos arquivos PDFs.", text_color="red")

from datetime import datetime

# Função para ser chamada quando o botão "Extrair CNPJ" for clicado
import os
from datetime import datetime

def extrair_cnpjs(label_erro, tipo):
    # Data atual no formato brasileiro (DD-MM-YYYY)
    data_atual = datetime.now().strftime("%d-%m-%Y")

    # Definindo o diretório e o nome do arquivo dependendo do tipo
    if tipo == 'homologados':
        pasta_vencedores = os.path.join(os.getcwd(), "vencedores_bll_compras_homologado")
        nome_arquivo_saida = f"dados_vencedores_homologados_bll_compras_{data_atual}.xlsx"
    elif tipo == 'adjudicados':
        pasta_vencedores = os.path.join(os.getcwd(), "vencedores_bll_compras_adjudicado")
        nome_arquivo_saida = f"dados_vencedores_adjudicados_bll_compras_{data_atual}.xlsx"
    else:
        label_erro.configure(text="Tipo inválido.", text_color="red")
        return

    # Verificando se a pasta existe e chamando a função para extrair os CNPJs
    if os.path.exists(pasta_vencedores):
        extrair_cnpjs_pasta(pasta_vencedores, nome_arquivo_saida, label_erro)
    else:
        label_erro.configure(text="Pasta de vencedores não encontrada!", text_color="red")


def criar_interface():
    def iniciar_busca():
        # Coleta os valores da interface
        data_inicio = entry_data_inicio.get()
        data_fim = entry_data_fim.get()
        status_processo = combo_status.get()
        extrair_bllcompras(data_inicio, data_fim, status_processo, label_contador_pdfs)

        
        # Validação simples 
        if not data_inicio or not data_fim or not status_processo:
            label_erro.configure(text="Preencha todos os campos!", text_color="red")
            return
        else:
            label_erro.configure(text="Executando busca...", text_color="green")
            root.update()

        # Executa o programa principal
        extrair_bllcompras(data_inicio, data_fim, status_processo,label_contador_pdfs)

       
    global root
    # Configuração da interface
    root = ctk.CTk()
    root.title("Busca BLL Compras")
    root.geometry("400x300")
    ctk.set_appearance_mode("Light")

    # Widgets
    label_titulo = ctk.CTkLabel(root, text="Buscar Processos BLL Compras", font=("Arial", 16, "bold"))
    label_titulo.pack(pady=10)

    # Campo para data inicial
    label_data_inicio = ctk.CTkLabel(root, text="Data Inicial (DD/MM/AAAA):")
    label_data_inicio.pack()
    entry_data_inicio = ctk.CTkEntry(root, placeholder_text="17/05/2024")
    entry_data_inicio.pack(pady=5)

    # Campo para data final
    label_data_fim = ctk.CTkLabel(root, text="Data Final (DD/MM/AAAA):")
    label_data_fim.pack()
    entry_data_fim = ctk.CTkEntry(root, placeholder_text="20/10/2024")
    entry_data_fim.pack(pady=5)

    # Dropdown para selecionar "Homologado" ou "Adjudicado"
    label_status = ctk.CTkLabel(root, text="Status do Processo:")
    label_status.pack()
    combo_status = ctk.CTkComboBox(root, values=["HOMOLOGADO", "ADJUDICADO"])
    combo_status.pack(pady=5)

    # Botão de executar
    btn_executar = ctk.CTkButton(root, text="Executar Busca", command=iniciar_busca)
    btn_executar.pack(pady=10)

    btn_extrair_cnpj_homologados = ctk.CTkButton(root, text="Extrair CNPJ Homologados", 
                                                 command=lambda: extrair_cnpjs(label_erro, 'homologados'))
    btn_extrair_cnpj_homologados.pack(pady=10)

    btn_extrair_cnpj_adjudicados = ctk.CTkButton(root, text="Extrair CNPJ Adjudicados", 
                                                 command=lambda: extrair_cnpjs(label_erro, 'adjudicados'))
    btn_extrair_cnpj_adjudicados.pack(pady=10)

    global label_contador_pdfs, label_contador_cnpjs
    label_contador_pdfs = ctk.CTkLabel(root, text="PDFs Baixados: 0")
    label_contador_pdfs.pack()

    label_contador_cnpjs = ctk.CTkLabel(root, text="CNPJs Extraídos: 0")
    label_contador_cnpjs.pack()

    # Label de erro ou sucesso
    label_erro = ctk.CTkLabel(root, text="")
    label_erro.pack()

    # Inicia a interface
    root.mainloop()





def iniciar_raspagem():
    threading.Thread(target=iniciar_raspagem_Portal_Nacional_de_Contratacoes_Publicas, daemon=True).start()




# Configuração do tema escuro
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("dark-blue")

# Criação da janela principal
app = ctk.CTk()
app.title("Sistema de Raspagem de Licitações")
screen_width = app.winfo_screenwidth()
screen_height = app.winfo_screenheight()
app.geometry(f"{screen_width}x{screen_height-40}+0+0")

# Título principal
title_label = ctk.CTkLabel(
    app, 
    text="Sistema de Raspagem de Licitações", 
    font=("Helvetica", 30, "bold"), 
    text_color="#ffffff"
)
title_label.pack(pady=20)

# Frame principal que contém todos os módulos
main_frame = ctk.CTkFrame(app, fg_color="transparent")
main_frame.pack(padx=20, pady=20, fill="both", expand=True)

# Frame para os 5 módulos
frame_modulos = ctk.CTkFrame(main_frame)
frame_modulos.pack(pady=30, fill="both", expand=True)

# Módulo 1 - Portal Nacional de Licitações
frame_pncp = ctk.CTkFrame(frame_modulos)
frame_pncp.pack(side="left", padx=20, pady=20, fill="y", expand=True)

# Imagem de exemplo para o módulo
img_pncp = ctk.CTkImage(Image.open("img.png"), size=(300, 100))  # Substitua pelo caminho correto da imagem
label_img_pncp = ctk.CTkLabel(frame_pncp, image=img_pncp, text="")
label_img_pncp.pack(pady=10)

label_pncp = ctk.CTkLabel(frame_pncp, text="Portal Nacional de Contratações Públicas", font=("Helvetica", 18, "bold"))
label_pncp.pack(pady=10)

button_pncp = ctk.CTkButton(
    frame_pncp, 
    text="Iniciar Raspagem", 
    command=criar_interface_pncp,
    font=("Helvetica", 14), 
    width=250, 
    height=40, 
    fg_color="#007acc",  
    hover_color="#005b99",  
)
button_pncp.pack(pady=10)

button_enviar_pncp = ctk.CTkButton(
    frame_pncp, 
    text="Enviar E-mails", 
    command=criar_interface_mautic,
    font=("Helvetica", 14), 
    width=250, 
    height=40, 
    fg_color="#4caf50",  
    hover_color="#388e3c",  
)
button_enviar_pncp.pack(pady=10)

# Módulo 2 - BLL - COMPRAS
frame_site1 = ctk.CTkFrame(frame_modulos)
frame_site1.pack(side="left", padx=20, pady=20, fill="y", expand=True)

# Imagem de exemplo para o módulo
img_site1 = ctk.CTkImage(Image.open("bll_compras.png"), size=(300, 100))  # Substitua pelo caminho correto da imagem
label_img_site1 = ctk.CTkLabel(frame_site1, image=img_site1, text="")
label_img_site1.pack(pady=10)

label_site1 = ctk.CTkLabel(frame_site1, text="BLL Compras", font=("Helvetica", 18, "bold"))
label_site1.pack(pady=10)

import openpyxl
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from tkinter import messagebox
from time import sleep
import random
import customtkinter as ctk
import re
from selenium.webdriver.common.action_chains import ActionChains

def verifica_ajudicacao(driver, root, label_status, cnpjs_set):
    situacao_texto = ""
    adjudicada_encontrada = False
    try:
        elementos_situacao = WebDriverWait(driver, 5).until(
            EC.visibility_of_all_elements_located((By.CSS_SELECTOR, "div[data-test='situacao-proposta']"))
        )

        for elemento_situacao in elementos_situacao:
            situacao_texto = elemento_situacao.text.strip()
            if situacao_texto.lower() == "adjudicada":  # Só pega o CNPJ se a situação for "adjudicada"
                adjudicada_encontrada = True
                print("A proposta está adjudicada. Realizando ações subsequentes...")

                try:
                    container = elemento_situacao.find_element(By.XPATH, "./ancestor::div[contains(@data-test, 'propostaItemEmSelecaoFornecedores')]")
                    container_cnpj = WebDriverWait(container, 10).until(
                        EC.visibility_of_element_located((By.XPATH, ".//span[@data-test='identificacao-participante']"))
                    )
                    cnpj = container_cnpj.text
                    print('CNPJ encontrado:', cnpj)

                    # Adiciona ao set para garantir que não haja duplicação
                    cnpjs_set.add(cnpj)

                    # Salva os CNPJs no Excel e atualiza o status
                    total_cnpjs = salvar_cnpjs_excel(cnpjs_set)
                    label_status.configure(text=f"Total de CNPJs encontrados: {total_cnpjs}")
                    print(f"Total de CNPJs encontrados: {total_cnpjs}")
                    root.update()

                except Exception as e:
                    print(f"Erro ao encontrar o botão de expansão: {e}")
                    continue
    except Exception as e:
        print(f"Erro ao buscar elementos de situação: {e}")
        return situacao_texto, False  # Caso não encontrar a situação

    return situacao_texto, adjudicada_encontrada

def raspar_pagina(driver, wb, root, label_status, cnpjs_set):
     # Coleta os links únicos
    elements = driver.find_elements(By.XPATH, "//i[@class='fa fa-tasks']")
    qtde_apps_card = len(driver.find_elements(By.XPATH, "//i[@class='fa fa-tasks']"))

    for index, element in enumerate(elements, start=0):
        

        elements = driver.find_elements(By.XPATH, "//i[@class='fa fa-tasks']")
        element = elements[index]
        #reencontra o elemento
        element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable(element)
        )

        url_inicial = driver.current_url

        print(f"Elemento {index+1} de {len(elements)}")
        driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", element)
        action = ActionChains(driver)
        action.move_to_element(element).click().perform()   
        print('clicado')  
        action.move_to_element(element).click().perform()  
        print('clicado dnv')
        print(qtde_apps_card)
        
        sleep(random.uniform(2, 3))

        url_atual = driver.current_url
        
        elements_details = driver.find_elements(By.XPATH, "//i[@class='fa-tasks fas']")

        if not elements_details and url_inicial != url_atual:
             print('Nao achei element details mas a URL mudou, voltando de pagina')
             driver.back()
             sleep(3)
             continue
        if elements_details and url_inicial != url_atual:
            print('Entrei em element details')
            for index_details, elementt in enumerate(elements_details, start=1):
                print(f"Clicando no elemento {index_details} de {len(elements_details)}")
                cont_element_details = len(elements_details)
                

                # Rolando suavemente até o elemento
                driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", elementt)
                

                # Ação para mover até o elemento e clicar
                action = ActionChains(driver)
                action.move_to_element(elementt).click().perform()

                print(f'Elemento {index} clicado com sucesso.')
                sleep(3)
                
                situacao_texto, adjudicada_encontrada = verifica_ajudicacao(driver, root, label_status, cnpjs_set)
                
                if not adjudicada_encontrada:
                        print('Nenhuma proposta adjudicada')
                        driver.back()
                        sleep(2)
                        driver.back()
                        sleep(3)
                else:
                    print('tem apenas 1 item, voltando pra pagina inicial')
                    driver.back()
                    sleep(2)
                    driver.back()  # Segunda vez
                    sleep(2)

                next_page = WebDriverWait(driver,5).until(
                    EC.element_to_be_clickable((By.XPATH,"//button[@class='p-ripple p-element p-paginator-next p-paginator-element p-link']"))
                )
                
                print(f'Elemento atual: {index}, último elemento: {qtde_apps_card - 1}')

                if index == qtde_apps_card - 1:
                    print(f'Elemento atual: {index}, último elemento: {(qtde_apps_card) - 1}')
                    next_page.click()
                    sleep(3)
                    raspar_pagina(driver, wb, root, label_status, cnpjs_set)
                else:
                    break
                    
                        
        else:
            print("Nenhum elemento encontrado.")
            continue

    wb.save("dados_vencedores_portal_compras_publicas.xlsx")
    sleep(15)
    driver.quit()
    messagebox.showinfo("Concluído", "Raspagem concluída com sucesso!")

def iniciar_raspagem_compras_gov(ano, root, label_status, cnpjs):
    # Inicializa o navegador com undetected_chromedriver
    driver = uc.Chrome()
    driver.get('https://cnetmobile.estaleiro.serpro.gov.br/comprasnet-web/public/compras')
    driver.maximize_window()
    sleep(random.uniform(1, 3))  # Delay aleatório

    # Criar a planilha para salvar os dados
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Link", "CNPJ", "Razão Social"])  

    # Interage com o filtro
    filtro = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, "//div[@class='p-radiobutton-box']"))
    )
    filtro.click()
    sleep(random.uniform(1, 3))  # Delay aleatório


    input_ano = WebDriverWait(driver,5).until(
        EC.element_to_be_clickable((By.XPATH,"//input[@placeholder='Ex: 102021']"))
    )
    input_ano.send_keys(ano)
    sleep(random.uniform(1, 3))

        # Clica no botão pesquisar
    button_pesquisar = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, "//button[@class='br-button is-primary']"))
    )
    button_pesquisar.click()
    sleep(random.uniform(1, 3))  # Delay aleatório

    raspar_pagina(driver, wb, root, label_status, cnpjs)

    wb.save("dados_vencedores_portal_compras_publicas.xlsx")
    sleep(5)
    driver.quit()
    messagebox.showinfo("Concluído", "Raspagem concluída com sucesso!")    


def salvar_cnpjs_excel(cnpjs_set):
    # Carregar o arquivo Excel existente ou criar um novo
    try:
        wb = openpyxl.load_workbook("dados_compras_gov.xlsx")
        ws = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CNPJs Extraídos"
        ws.append([None, "CNPJ"])  # Cabeçalho do arquivo Excel

    # Adiciona os CNPJs do set (sem duplicados)
    for cnpj in cnpjs_set:
        # Verifica se o CNPJ já está na planilha antes de adicionar
        cell_values = [cell.value for cell in ws['B']]  # Lista de valores existentes na coluna B
        if cnpj not in cell_values:
            ws.append([None, cnpj])  # Adiciona o CNPJ apenas se não estiver na planilha

    from datetime import datetime
    # Salva o arquivo
    data_atual = datetime.now().strftime("%d-%m-%Y")
    nome_arquivo = f"dados_compras_gov_{data_atual}.xlsx"
    wb.save(nome_arquivo)
    print(f"Total de CNPJs únicos salvos no arquivo Excel: {len(cnpjs_set)}")

    return len(cnpjs_set)  # Retorna a quantidade de CNPJs únicos adicionados

def criar_interface_compras_gov():
    ctk.set_appearance_mode("dark")
    ctk.set_default_color_theme("blue")

    root = ctk.CTk()
    root.geometry("400x300")
    root.title("Raspagem Compras Gov")

    ctk.CTkLabel(root, text="Digite o ano (Ex: 2023):").pack(pady=10)
    entry_ano = ctk.CTkEntry(root, placeholder_text="Digite o ano")
    entry_ano.pack(pady=10)

    label_status = ctk.CTkLabel(root, text="Status: Aguardando ação...", font=("Arial", 14))
    label_status.pack(pady=10)

    cnpjs_set = set()  # Inicializa o set de CNPJs aqui

    def iniciar():
        ano = entry_ano.get()
        if ano:
            iniciar_raspagem_compras_gov(ano, root, label_status, cnpjs_set)
        else:
            messagebox.showwarning("Erro", "Por favor, insira um ano.")

    ctk.CTkButton(root, text="Iniciar Raspagem", command=iniciar).pack(pady=10)
    root.mainloop()

import openpyxl
import customtkinter as ctk
from tkinter import messagebox
from tkinter.filedialog import askopenfilenames  # Para abrir a caixa de seleção de arquivos
import os
import glob
from datetime import datetime

# Função para carregar os CNPJs de arquivos
def carregar_cnpjs_de_arquivos(arquivos):
    cnpjs_unicos = set()  # Usando um set para garantir que não haja CNPJs duplicados
    cnpjs_ocorrencias = []  # Lista para armazenar todas as ocorrências de CNPJs (mesmo duplicados)

    for arquivo in arquivos:
        wb = openpyxl.load_workbook(arquivo)
        ws = wb.active
        
        # Itera pelas linhas e coleta os CNPJs
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):  # Pula o cabeçalho
            cnpj = row[1]
            if cnpj:
                cnpj = cnpj.strip()
                cnpjs_ocorrencias.append(cnpj)  # Adiciona todas as ocorrências
                cnpjs_unicos.add(cnpj)  # Adiciona apenas CNPJs únicos

    return cnpjs_unicos, cnpjs_ocorrencias

# Função para salvar os CNPJs e suas ocorrências em um arquivo único
def salvar_cnpjs_unicos(cnpjs_unicos, cnpjs_ocorrencias):
    # Obtém a data no formato brasileiro (DD-MM-AAAA)
    data_hoje = datetime.now().strftime("%d-%m-%Y")
    
    # Nome do arquivo onde os CNPJs serão salvos
    nome_arquivo = f"cnpjs_unicos_{data_hoje}.xlsx"

    # Criar ou carregar o arquivo Excel onde os dados serão armazenados
    if os.path.exists(nome_arquivo):
        wb = openpyxl.load_workbook(nome_arquivo)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["CNPJ", "Ocorrências"])  # Cabeçalho da planilha

    # Adiciona os CNPJs e suas ocorrências na planilha
    for cnpj in cnpjs_unicos:
        ocorrencias = cnpjs_ocorrencias.count(cnpj)  # Conta quantas vezes o CNPJ aparece
        ws.append([cnpj, ocorrencias])  # Adiciona o CNPJ e o número de ocorrências

    wb.save(nome_arquivo)
    print(f"CNPJs únicos e suas ocorrências salvos em: {nome_arquivo}")

# Função para renomear os arquivos para indicar que foram processados
def renomear_arquivo_processado(arquivos):
    data_hoje = datetime.now().strftime("%d-%m-%Y")

    for arquivo in arquivos:
        # Renomeia o arquivo original para algo como: cnpjs_processados_{data}.xlsx
        novo_nome = f"cnpjs_processados_{data_hoje}_{os.path.basename(arquivo)}"
        os.rename(arquivo, novo_nome)
        print(f"Arquivo renomeado para: {novo_nome}")

# Função para iniciar o processo de eliminação de CNPJs duplicados
def eliminar_cnpjs():
    # Abre uma caixa de diálogo para o usuário selecionar os arquivos
    arquivos = askopenfilenames(
        title="Selecione os arquivos de dados", 
        filetypes=[("Arquivos Excel", "*.xlsx")], 
        initialdir=os.getcwd()  # Inicia na pasta atual
    )
    
    # Verifica se o usuário selecionou algum arquivo
    if not arquivos:
        messagebox.showwarning("Atenção", "Nenhum arquivo foi selecionado.")
        return

    # Carregar CNPJs e ocorrências de todos os arquivos selecionados
    cnpjs_unicos, cnpjs_ocorrencias = carregar_cnpjs_de_arquivos(arquivos)

    # Salvar todos os CNPJs na mesma planilha
    salvar_cnpjs_unicos(cnpjs_unicos, cnpjs_ocorrencias)

    # Renomeia os arquivos processados
    renomear_arquivo_processado(arquivos)

    messagebox.showinfo("Concluído", "CNPJs duplicados eliminados com sucesso!")

# Função para criar a interface gráfica com Tkinter
def criar_interface_eliminar_cnpj():
    ctk.set_appearance_mode("dark")
    ctk.set_default_color_theme("blue")

    root = ctk.CTk()
    root.geometry("400x300")
    root.title("Eliminação de CNPJs Duplicados")

    ctk.CTkButton(root, text="Selecionar e Processar Arquivos", command=eliminar_cnpjs).pack(pady=10)

    root.mainloop()


import os
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
import customtkinter as ctk
import glob  # Para buscar arquivos com padrão
from tkinter import messagebox

def criar_interface_raspagem_emails():
    janela = ctk.CTk()  # Janela principal
    janela.geometry("500x300")
    janela.title("Raspagem de E-mails")

    # Label para exibir o número de e-mails raspados
    label_contador = ctk.CTkLabel(janela, text="E-mails raspados: 0", font=("Arial", 16))
    label_contador.pack(pady=10)

    # Combobox para selecionar o site
    label_selecione = ctk.CTkLabel(janela, text="Selecione o site:", font=("Arial", 14))
    label_selecione.pack(pady=5)

    site_selecionado = ctk.StringVar(value="CNPJ Já")
    dropdown_site = ctk.CTkComboBox(janela, values=["Consulta Guru", "CNPJ Já"], variable=site_selecionado)
    dropdown_site.pack(pady=5)

    # Função de depuração para verificar a seleção
    def verificar_selecao():
        print(f"Site selecionado: {site_selecionado.get()}")

    # Botão para iniciar a raspagem
    def iniciar_raspagem():
        site = site_selecionado.get()
        print(f"Valor do site selecionado no botão: {site}")
        if site == "Consulta Guru":
            consulta_cnpj_gratis(label_contador, janela)
        elif site == "CNPJ Já":
            consulta_cnpj_ja(label_contador, janela)

    # Botão de depuração para verificar o valor da seleção
    botao_verificar = ctk.CTkButton(janela, text="Verificar Seleção", command=verificar_selecao)
    botao_verificar.pack(pady=5)

    botao_iniciar = ctk.CTkButton(janela, text="Iniciar Raspagem", command=iniciar_raspagem)
    botao_iniciar.pack(pady=20)

    janela.mainloop()

def salvar_emails(resultados):
    """
    Salva os resultados no arquivo de e-mails, criando novos arquivos se o número de linhas exceder 2000.
    """
    nome_arquivo_base = "emails_vencedores"
    arquivos_existentes = glob.glob(f"{nome_arquivo_base}_*.xlsx")
    indice_arquivo = len(arquivos_existentes) + 1

    if not arquivos_existentes:
        nome_arquivo = f"{nome_arquivo_base}.xlsx"
    else:
        nome_arquivo = f"{nome_arquivo_base}_{indice_arquivo}.xlsx"

    df_resultados = pd.DataFrame(resultados)

    if os.path.exists(nome_arquivo):
        df_existente = pd.read_excel(nome_arquivo)
        if len(df_existente) >= 2000:
            indice_arquivo += 1
            nome_arquivo = f"{nome_arquivo_base}_{indice_arquivo}.xlsx"

    df_resultados.to_excel(nome_arquivo, index=False)
    print(f"E-mails salvos no arquivo '{nome_arquivo}'.")

def consulta_cnpj_gratis(label_contador, janela):
    arquivos = glob.glob("cnpjs_unicos*.xlsx")
    
    if not arquivos:
        messagebox.showwarning("Atenção", "O arquivo 'cnpjs_unicos.xlsx' não foi encontrado! Elimine os CNPJ duplicados!")
        return

    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

    def limpar_cnpj(cnpj):
        return ''.join(filter(str.isdigit, str(cnpj)))

    total_emails = 0
    resultados = []
    emails_unicos = set()

    try:
        for arquivo in arquivos:
            if os.path.exists(arquivo):
                df = pd.read_excel(arquivo)

                for cnpj in df.iloc[:, 0]:  # Segunda coluna
                    cnpj_limpo = limpar_cnpj(cnpj)
                    if cnpj_limpo:
                        link = f"https://consulta.guru/consultar-cnpj-gratis/{cnpj_limpo}"
                        driver.get(link)
                        sleep(5)

                        try:
                            email_element = driver.find_element(By.XPATH, '//p[contains(text(), "@")]')
                            email = email_element.text.strip()

                            # Capturando a razão social
                            razao_social_element = driver.find_element(By.XPATH, '//div[@id="overview"]//h1')
                            razao_social = razao_social_element.text.strip()

                            if email not in emails_unicos:
                                emails_unicos.add(email)
                                resultados.append({"Razao Social": razao_social, "Email": email, "CNPJ": cnpj_limpo})
                                total_emails += 1
                                label_contador.configure(text=f"E-mails raspados: {total_emails}")
                                janela.update()
                        except Exception:
                            print(f"Erro ao encontrar o e-mail ou razão social para o CNPJ: {cnpj_limpo}")
                            pass  # Se não achar, continua

                        salvar_emails(resultados)
                        sleep(12)

        salvar_emails(resultados)
    finally:
        driver.quit()

def consulta_cnpj_ja(label_contador, janela):
    arquivos = glob.glob("cnpjs_unicos*.xlsx")

    if not arquivos:
        messagebox.showwarning("Atenção", "O arquivo 'cnpjs_unicos.xlsx' não foi encontrado! Elimine os CNPJ duplicados!")
        return

    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

    def limpar_cnpj(cnpj):
        return ''.join(filter(str.isdigit, str(cnpj)))

    total_emails = 0
    resultados = []
    emails_unicos = set()

    try:
        for arquivo in arquivos:
            if os.path.exists(arquivo):
                df = pd.read_excel(arquivo)

                for cnpj in df.iloc[:, 0]:  # Segunda coluna
                    cnpj_limpo = limpar_cnpj(cnpj)
                    if cnpj_limpo:
                        link = f"https://cnpja.com/office/{cnpj_limpo}"
                        driver.get(link)
                        sleep(3)

                        try:
                            email_element = WebDriverWait(driver, 10).until(
                                EC.presence_of_element_located((By.XPATH, "//span[contains(text(), '@')]"))
                            )
                            email = email_element.text.strip()

                            # Capturando a razão social
                            razao_social_element = driver.find_element(By.XPATH, '//div[@class="inline cursor-copy "]/span')
                            razao_social = razao_social_element.text.strip()

                            if email not in emails_unicos:
                                if email != 'contato@cnpja.com':
                                    emails_unicos.add(email)
                                    resultados.append({"Razao Social": razao_social, "Email": email, "CNPJ": cnpj_limpo})
                                    total_emails += 1
                                    label_contador.configure(text=f"E-mails raspados: {total_emails}")
                                    janela.update()
                        except Exception:
                            print(f"Erro ao encontrar o e-mail ou razão social para o CNPJ: {cnpj_limpo}")
                            pass  # Se não achar, continua

                        salvar_emails(resultados)
                        sleep(12)

        salvar_emails(resultados)
    finally:
        driver.quit()


button_site1 = ctk.CTkButton(
    frame_site1, 
    text="Iniciar Raspagem", 
    command=criar_interface,
    font=("Helvetica", 14), 
    width=250, 
    height=40, 
    fg_color="#007acc",  
    hover_color="#005b99",  
)
button_site1.pack(pady=10)

button_enviar_site1 = ctk.CTkButton(
    frame_site1, 
    text="Enviar E-mails", 
    command=criar_interface_mautic,
    font=("Helvetica", 14), 
    width=250, 
    height=40, 
    fg_color="#4caf50",  
    hover_color="#388e3c",  
)
button_enviar_site1.pack(pady=10)

# Módulo 3 - Outro Site de Licitações
frame_site2 = ctk.CTkFrame(frame_modulos)
frame_site2.pack(side="left", padx=20, pady=20, fill="y", expand=True)

# Imagem de exemplo para o módulo
img_site2 = ctk.CTkImage(Image.open("compras_gov.png"), size=(300, 100))  # Substitua pelo caminho correto da imagem
label_img_site2 = ctk.CTkLabel(frame_site2, image=img_site2, text="")
label_img_site2.pack(pady=10)

label_site2 = ctk.CTkLabel(frame_site2, text="Compras GOV", font=("Helvetica", 18, "bold"))
label_site2.pack(pady=10)
#funcionando

button_site2 = ctk.CTkButton(
    frame_site2, 
    text="Iniciar Raspagem", 
    command=criar_interface_compras_gov,
    font=("Helvetica", 14), 
    width=250, 
    height=40, 
    fg_color="#007acc",  
    hover_color="#005b99",  
)
button_site2.pack(pady=10)

button_enviar_site2 = ctk.CTkButton(
    frame_site2, 
    text="Enviar E-mails", 
    command=criar_interface_mautic,
    font=("Helvetica", 14), 
    width=250, 
    height=40, 
    fg_color="#4caf50",  
    hover_color="#388e3c",  
)
button_enviar_site2.pack(pady=10)

# Rodapé
footer_frame = ctk.CTkFrame(app, fg_color="transparent")
footer_frame.pack(side="bottom", pady=20)

button_eliminar = ctk.CTkButton(
    footer_frame, 
    text="Eliminar CNPJ Repetido", 
    command=criar_interface_eliminar_cnpj,
    font=("Helvetica", 14), 
    width=250, 
    height=40, 
    fg_color="#ff5722",  
    hover_color="#d43f00",  
)
button_eliminar.pack(pady=10)

button_raspagem_emails = ctk.CTkButton(
    footer_frame, 
    text="Raspagem de E-mails", 
    command=criar_interface_raspagem_emails,
    font=("Helvetica", 14), 
    width=250, 
    height=40, 
    fg_color="#007acc",  
    hover_color="#005b99",  
)
button_raspagem_emails.pack(side="left", padx=10)


# Iniciar o loop principal da interface gráfica
app.mainloop()
#app