import customtkinter as ctk
from tkinter import messagebox
from PIL import Image
import openpyxl
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import time
import pyautogui
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import StaleElementReferenceException, ElementClickInterceptedException
from selenium.webdriver.common.action_chains import ActionChains
import os
import re
import pdfplumber
import pandas as pd
import customtkinter as ctk
import json
import threading
from selenium.common.exceptions import StaleElementReferenceException
from selenium.common.exceptions import NoSuchElementException
from time import sleep
import tkinter as tk
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

# Funções de raspagem (essas são apenas placeholders, substitua pelas suas funções de raspagem reais)
def iniciar_raspagem_pncp():
    messagebox.showinfo("Raspagem Portal Nacional", "Iniciando raspagem de dados no Portal Nacional de Licitações...")

PROGRESSO_JSON = "progresso.json"

def salvar_progresso(pdfs_baixados, cnpjs_extraidos):
    progresso = {
        "pdfs_baixados": pdfs_baixados,
        "cnpjs_extraidos": cnpjs_extraidos
    }
    with open(PROGRESSO_JSON, "w") as file:
        json.dump(progresso, file)

# Função para carregar progresso
def carregar_progresso():
    if os.path.exists(PROGRESSO_JSON):
        with open(PROGRESSO_JSON, "r") as file:
            return json.load(file)
    return {"pdfs_baixados": 0, "cnpjs_extraidos": 0}

def limpa_campo():
    for _ in range(10):
        pyautogui.press('right')
    for _ in range(10):
        pyautogui.press('backspace')

def scroll_ate_resultados_estabilizarem(driver):
    # Obter o número inicial de resultados
    resultado_atual = driver.find_element(By.ID, "footResultCount").text
    resultado_atual = int(resultado_atual.split(":")[-1].strip())

    while True:
        # Rolar para baixo
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(4)  # Aguarda o carregamento dos novos resultados

        # Obter o novo número de resultados
        novo_resultado = driver.find_element(By.ID, "footResultCount").text
        novo_resultado = int(novo_resultado.split(":")[-1].strip())

        # Verificar se o número de resultados mudou
        if novo_resultado == resultado_atual:
            print(f"Número de resultados estabilizado: {novo_resultado}")
            break
        else:
            resultado_atual = novo_resultado  # Atualiza o número de resultados

def extrair_bllcompras(data_inicio, data_fim, status_processo,label_contador_pdfs):

    progresso_json = "progresso.json"

    if os.path.exists(progresso_json):
        os.remove(progresso_json)
        print(f"Arquivo {progresso_json} excluído.")

    current_dir = os.getcwd()
    if status_processo == "HOMOLOGADO":
        pasta_destino = os.path.join(current_dir, "vencedores_bll_compras_homologado")
    elif status_processo == "ADJUDICADO":
        pasta_destino = os.path.join(current_dir, "vencedores_bll_compras_adjudicado")
    else:
        print(f"Erro: Status de processo inválido: {status_processo}")
        return

    # Verifica se a pasta existe, se não, cria a pasta
    if not os.path.exists(pasta_destino):
        os.makedirs(pasta_destino)
        print(f"Pasta criada: {pasta_destino}")
    chrome_options = webdriver.ChromeOptions()
    prefs = {
        "download.default_directory": pasta_destino,  # Define o diretório atual
        "download.prompt_for_download": False,      # Não perguntar antes de baixar
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True                # Ativa o download seguro
    }
    chrome_options.add_experimental_option("prefs", prefs)

    driver = webdriver.Chrome(options=chrome_options)
    driver.get('https://bllcompras.com/Process/ProcessSearchPublic?param1=0#')
    driver.maximize_window()

    

    # Aguarda até que o elemento de seleção de status esteja presente
    select_element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "fkStatus"))
    )
    select = Select(select_element)
    select.select_by_visible_text(status_processo)

    # Aguarda até que o botão de busca avançada seja clicável
    busca_avancada = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.XPATH, "//a[@title='BUSCA AVANÇADA']"))
    )
    busca_avancada.click()

    public_inicio = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.XPATH, "//input[@data-target='#DateStart']"))
    )
    public_inicio.click()
    public_inicio.clear()
    public_inicio.send_keys(data_inicio)

    # Preencher a data final
    public_final = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.XPATH, "//input[@data-target='#DateEnd']"))
    )
    public_final.click()
    public_final.clear()
    public_final.send_keys(data_fim)

    # Aguarda até que o ícone de pesquisa seja clicável
    icon = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, "button i.fas.fa-search"))
    )
    icon.click()

    scroll_ate_resultados_estabilizarem(driver)

    # Espera até que a lista de elementos de informações esteja carregada
    WebDriverWait(driver, 10).until(
        EC.presence_of_all_elements_located((By.CSS_SELECTOR, "i.fas.fa-info-circle"))
    )

    original_window = driver.current_window_handle
    original_url = driver.current_url

    elements = driver.find_elements(By.CSS_SELECTOR, "i.fas.fa-info-circle")

    pdfs_baixados = 0
    
    progresso = carregar_progresso()
    pdfs_baixados = progresso["pdfs_baixados"]

    for index, element in enumerate(elements, start=1):
        print(f"Elemento {index} de {len(elements)}")

        elements = driver.find_elements(By.CSS_SELECTOR, "i.fas.fa-info-circle")
        element = elements[index]
        
        # Reencontrando o elemento antes de clicar para garantir que ele ainda é válido
        element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable(element)
        )
        driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", element)
        action = ActionChains(driver)
        action.move_to_element(element).click().perform()
        
        time.sleep(2)
        print("BOTAO 1 CLICADO")
        

        WebDriverWait(driver, 10).until(
        lambda driver: len(driver.window_handles) > 1  # Garantir que uma nova janela foi aberta
        )
        
        # Trocar para a nova janela
        new_window = [window for window in driver.window_handles if window != original_window][0]
        driver.switch_to.window(new_window)

        current_url = driver.current_url
        print("URL da nova janela:", current_url)

        try:
            # Exemplo: esperando por um botão que deve aparecer na nova janela
            relatorios_button = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, "//button[@title='Relatórios']"))
            )
            print("Elemento 'Relatórios' apareceu, a página foi carregada corretamente.")
            
            # Clicar no botão 'Relatórios' caso apareça
            relatorios_button.click()
            time.sleep(2)
            
        except Exception as e:
            print("O elemento 'Relatórios' não apareceu ou houve um erro:", e)
        
        try:
            # Aguarda a presença do elemento <b> e verifica o texto
            mensagem_elemento = WebDriverWait(driver, 3).until(
                EC.presence_of_element_located((By.XPATH, "//b"))
            )
            mensagem_texto = mensagem_elemento.text

            # Verifica o conteúdo do texto
            if "relatórios ainda não estão disponíveis" in mensagem_texto:
                print("Relatórios não estão disponíveis. Pulando para o próximo item...")
                pyautogui.hotkey('ctrl','w')
                driver.switch_to.window(original_window)
                continue  # Passa para o próximo elemento da lista
        except Exception:
            print("Elemento <b> não encontrado. Continuando o download...")

        WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.TAG_NAME, "tr"))
        )

        linhas = driver.find_elements(By.TAG_NAME, "tr")
    

        for linha in linhas:
            try:
                # Verifica se a linha contém o texto "VENCEDORES DO PROCESSO"
                if "VENCEDORES DO PROCESSO" in linha.text.upper():
                    print("Linha encontrada com 'VENCEDORES DO PROCESSO'.")

                    # Procura todas as tags <a> com o atributo 'download' dentro da linha
                    links_com_download = linha.find_elements(By.CSS_SELECTOR, "a[download]")
                    print(f"Total de links com 'download' encontrados: {len(links_com_download)}")

                    # Itera sobre os links com atributo 'download'
                    for link in links_com_download:
                        # Verifica se o atributo 'download' ou texto contém "VENCEDORES DO PROCESSO"
                        if "VencedoresProcesso" in link.get_attribute("download"):
                            print("Link correto encontrado com 'VencedoresProcesso' no atributo 'download'.")

                            try:
                                botao_download = link.find_element(By.CSS_SELECTOR, "i.fas.fa-download")
                                print("Botão de download encontrado. Tentando clicar...")

                                # Realiza o scroll até o botão para garantir que ele esteja visível
                                driver.execute_script("arguments[0].scrollIntoView(true);", botao_download)
                                time.sleep(1)

                                # Realiza o clique com ActionChains
                                ActionChains(driver).move_to_element(botao_download).click().perform()
                                print("Download iniciado com sucesso.")
                                pdfs_baixados += 1
                                label_contador_pdfs.configure(text=f"PDFs    Baixados: {pdfs_baixados}")
                                salvar_progresso(pdfs_baixados, progresso["cnpjs_extraidos"])
                                root.update()
                                time.sleep(5)  
                                download_realizado = True
                                break  

                            except Exception as e:
                                print(f"Erro ao tentar clicar no ícone de download: {e}")
                        if download_realizado:
                            print("Download concluído para esta linha. Encerrando busca na linha.")
                            break  
                        else:
                            print("Link ignorado, não corresponde ao 'VencedoresProcesso'.")

            except Exception as e:
                print(f"Erro ao processar a linha: {e}")

        # Fecha a aba atual e retorna à aba original
        pyautogui.hotkey('ctrl', 'w')
        driver.switch_to.window(original_window)

    # Espera um pouco mais para garantir que o download seja iniciado
    time.sleep(5)  # Pausa de 5 segundos antes de fechar o driver
    scroll_ate_resultados_estabilizarem(driver)
    driver.quit()  # Fechamento do driver após a execução

def extrair_cnpjs_pasta(pasta, nome_arquivo_saida, label_erro):
    """
    Extrai CNPJs de todos os PDFs em uma pasta específica e salva em um arquivo Excel.
    
    :param pasta: Caminho da pasta contendo os PDFs
    :param nome_arquivo_saida: Nome do arquivo Excel de saída onde os CNPJs serão salvos
    :param label_erro: Widget de erro para exibir mensagens na interface
    """
    # Padrão regex para capturar CNPJs
    padrao_cnpj = r'\b\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}\b'
    progresso = carregar_progresso()
    cnpjs_encontrados = set()
    total_cnpjs = progresso["cnpjs_extraidos"]

    # Itera sobre os arquivos na pasta
    for arquivo in os.listdir(pasta):
        if arquivo.endswith(".pdf"):
            caminho_arquivo = os.path.join(pasta, arquivo)
            print(f"Processando: {caminho_arquivo}")

            # Abre o PDF e extrai texto
            try:
                with pdfplumber.open(caminho_arquivo) as pdf:
                    for pagina in pdf.pages:
                        texto = pagina.extract_text()
                        if texto:
                            # Procura por CNPJs no texto
                            cnpjs = re.findall(padrao_cnpj, texto)
                            cnpjs_encontrados.update(cnpjs)
                            total_cnpjs += len(cnpjs)
                            label_contador_cnpjs.configure(text=f"CNPJs Extraídos: {total_cnpjs}")
                            salvar_progresso(progresso["pdfs_baixados"], total_cnpjs)
                            root.update()
                            if cnpjs:
                                print(f"CNPJs encontrados no arquivo {arquivo}: {cnpjs}")
                                cnpjs_encontrados.update(cnpjs)
            except Exception as e:
                print(f"Erro ao processar {arquivo}: {e}")

    # Remove duplicatas de CNPJs
    cnpjs_unicos = list(set(cnpjs_encontrados))

    # Exibe todos os CNPJs encontrados
    if cnpjs_unicos:
        print("\nCNPJs encontrados:")
        for cnpj in cnpjs_unicos:
            print(cnpj)
        
        # Salva os CNPJs em um arquivo Excel
        caminho_saida = os.path.join(os.getcwd(), nome_arquivo_saida)  # Caminho completo para o arquivo
        df = pd.DataFrame({
            "Índice": range(1, len(cnpjs_unicos) + 1),
            "CNPJs": cnpjs_unicos  # Insira os CNPJs na segunda coluna
        })
        df.to_excel(caminho_saida, index=False)
        print(f"\nCNPJs salvos em: {caminho_saida}")
        
        # Atualiza a interface com uma mensagem de sucesso
        label_erro.configure(text=f"CNPJs extraídos e salvos em {nome_arquivo_saida}", text_color="green")
    else:
        label_erro.configure(text="Nenhum CNPJ encontrado nos arquivos PDFs.", text_color="red")


# Função para ser chamada quando o botão "Extrair CNPJ" for clicado
def extrair_cnpjs(label_erro, tipo):
    if tipo == 'homologados':
        pasta_vencedores = os.path.join(os.getcwd(), "vencedores_bll_compras_homologado")
        nome_arquivo_saida = "dados_vencedores_homologados_bll_compras.xlsx"
    elif tipo == 'adjudicados':
        pasta_vencedores = os.path.join(os.getcwd(), "vencedores_bll_compras_adjudicado")
        nome_arquivo_saida = "dados_vencedores_adjudicados_bll_compras.xlsx"
    else:
        label_erro.configure(text="Tipo inválido.", text_color="red")
        return

    if os.path.exists(pasta_vencedores):
        extrair_cnpjs_pasta(pasta_vencedores, nome_arquivo_saida, label_erro)
    else:
        label_erro.configure(text="Pasta de vencedores não encontrada!", text_color="red")

def criar_interface():
    def iniciar_busca():
        # Coleta os valores da interface
        data_inicio = entry_data_inicio.get()
        data_fim = entry_data_fim.get()
        status_processo = combo_status.get()
        extrair_bllcompras(data_inicio, data_fim, status_processo, label_contador_pdfs)

        
        # Validação simples 
        if not data_inicio or not data_fim or not status_processo:
            label_erro.configure(text="Preencha todos os campos!", text_color="red")
            return
        else:
            label_erro.configure(text="Executando busca...", text_color="green")
            root.update()

        # Executa o programa principal
        extrair_bllcompras(data_inicio, data_fim, status_processo,label_contador_pdfs)

       
    global root
    # Configuração da interface
    root = ctk.CTk()
    root.title("Busca BLL Compras")
    root.geometry("400x300")
    ctk.set_appearance_mode("Light")

    # Widgets
    label_titulo = ctk.CTkLabel(root, text="Buscar Processos BLL Compras", font=("Arial", 16, "bold"))
    label_titulo.pack(pady=10)

    # Campo para data inicial
    label_data_inicio = ctk.CTkLabel(root, text="Data Inicial (DD/MM/AAAA):")
    label_data_inicio.pack()
    entry_data_inicio = ctk.CTkEntry(root, placeholder_text="17/05/2024")
    entry_data_inicio.pack(pady=5)

    # Campo para data final
    label_data_fim = ctk.CTkLabel(root, text="Data Final (DD/MM/AAAA):")
    label_data_fim.pack()
    entry_data_fim = ctk.CTkEntry(root, placeholder_text="20/10/2024")
    entry_data_fim.pack(pady=5)

    # Dropdown para selecionar "Homologado" ou "Adjudicado"
    label_status = ctk.CTkLabel(root, text="Status do Processo:")
    label_status.pack()
    combo_status = ctk.CTkComboBox(root, values=["HOMOLOGADO", "ADJUDICADO"])
    combo_status.pack(pady=5)

    # Botão de executar
    btn_executar = ctk.CTkButton(root, text="Executar Busca", command=iniciar_busca)
    btn_executar.pack(pady=10)

    btn_extrair_cnpj_homologados = ctk.CTkButton(root, text="Extrair CNPJ Homologados", 
                                                 command=lambda: extrair_cnpjs(label_erro, 'homologados'))
    btn_extrair_cnpj_homologados.pack(pady=10)

    btn_extrair_cnpj_adjudicados = ctk.CTkButton(root, text="Extrair CNPJ Adjudicados", 
                                                 command=lambda: extrair_cnpjs(label_erro, 'adjudicados'))
    btn_extrair_cnpj_adjudicados.pack(pady=10)

    global label_contador_pdfs, label_contador_cnpjs
    label_contador_pdfs = ctk.CTkLabel(root, text="PDFs Baixados: 0")
    label_contador_pdfs.pack()

    label_contador_cnpjs = ctk.CTkLabel(root, text="CNPJs Extraídos: 0")
    label_contador_cnpjs.pack()

    # Label de erro ou sucesso
    label_erro = ctk.CTkLabel(root, text="")
    label_erro.pack()

    # Inicia a interface
    root.mainloop()


def iniciar_raspagem_site2():
    messagebox.showinfo("Raspagem Site 2", "Iniciando raspagem de dados no Site 2...")

def iniciar_raspagem_site3():
    messagebox.showinfo("Raspagem Site 3", "Iniciando raspagem de dados no Site 3...")

def iniciar_raspagem_site4():
    messagebox.showinfo("Raspagem Site 4", "Iniciando raspagem de dados no Site 4...")

import openpyxl
import customtkinter as ctk
from tkinter import messagebox
import os

def carregar_cnpjs_de_arquivos(arquivos):
    cnpjs_unicos = set()  # Usando um set para garantir que não haja CNPJs duplicados
    cnpjs_ocorrencias = []  # Lista para armazenar todas as ocorrências de CNPJs (mesmo duplicados)

    for arquivo in arquivos:
        wb = openpyxl.load_workbook(arquivo)
        ws = wb.active
        
        # Itera pelas linhas e coleta os CNPJs
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):  # Pula o cabeçalho
            cnpj = row[1]
            if cnpj:
                cnpj = cnpj.strip()
                cnpjs_ocorrencias.append(cnpj)  # Adiciona todas as ocorrências
                cnpjs_unicos.add(cnpj)  # Adiciona apenas CNPJs únicos

    return cnpjs_unicos, cnpjs_ocorrencias

def salvar_cnpjs_unicos(cnpjs_unicos, cnpjs_ocorrencias, nome_arquivo="cnpjs_unicos.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["CNPJ", "Ocorrências"])  # Cabeçalho da planilha

    for cnpj in cnpjs_unicos:
        ocorrencias = cnpjs_ocorrencias.count(cnpj)  # Conta quantas vezes o CNPJ aparece
        ws.append([cnpj, ocorrencias])  # Adiciona o CNPJ e o número de ocorrências

    wb.save(nome_arquivo)
    print(f"CNPJs únicos e suas ocorrências salvos em: {nome_arquivo}")

def eliminar_cnpjs():
    # Lista de arquivos que devem ser processados
    arquivos = [
        "dados_fornecedores.xlsx", 
        "dados_vencedores_diario_sp.xlsx", 
        "dados_vencedores_homologados_bll_compras.xlsx", 
        "dados_vencedores_adjudicados_bll_compras.xlsx", 
        "dados_fornecedores_5.xlsx"
    ]

    # Verifica se os arquivos existem
    arquivos_existentes = [arquivo for arquivo in arquivos if os.path.exists(arquivo)]

    if not arquivos_existentes:
        # Exibe mensagem de aviso e retorna
        messagebox.showwarning("Atenção", "Nenhum arquivo de dados foi encontrado para eliminar CNPJs duplicados.")
        return

    cnpjs_unicos, cnpjs_ocorrencias = carregar_cnpjs_de_arquivos(arquivos_existentes)

    salvar_cnpjs_unicos(cnpjs_unicos, cnpjs_ocorrencias)
    messagebox.showinfo("Concluído", "CNPJs duplicados eliminados com sucesso!")


def enviar_emails_site1():
    messagebox.showinfo("Enviar E-mails Site 1", "Enviando e-mails para o Site 1...")

def enviar_emails_site2():
    messagebox.showinfo("Enviar E-mails Site 2", "Enviando e-mails para o Site 2...")

def enviar_emails_site3():
    messagebox.showinfo("Enviar E-mails Site 3", "Enviando e-mails para o Site 3...")

def enviar_emails_site4():
    messagebox.showinfo("Enviar E-mails Site 4", "Enviando e-mails para o Site 4...")

def iniciar_raspagem_compras_gov(dia_inicio, mes_inicio, ano_inicio, dia_fim, mes_fim, ano_fim, cnpj_label, janela):
    driver = webdriver.Chrome()
    driver.get('https://www.imprensaoficial.com.br/ENegocios/BuscaENegocios_14_1.aspx#12/12/2024')
    sleep(2)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Link", "CNPJ", "Razão Social"])

    # Preencher Status ENCERRADA
    encerrada = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.XPATH, "//select[@id='content_content_content_Status_cboStatus']/option[text()='ENCERRADA']"))
    )
    encerrada.click()

    # Preencher Data Inicial
    xpath_dia_inicio = f"//select[@id='content_content_content_Status_cboAberturaSecaoInicioDia']/option[@value='{dia_inicio}']"
    xpath_mes_inicio = f"//select[@id='content_content_content_Status_cboAberturaSecaoInicioMes']/option[@value='{mes_inicio}']"
    xpath_ano_inicio = f"//select[@id='content_content_content_Status_cboAberturaSecaoInicioAno']/option[@value='{ano_inicio}']"

    WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, xpath_dia_inicio))).click()
    WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, xpath_mes_inicio))).click()
    WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, xpath_ano_inicio))).click()

    # Preencher Data Final
    xpath_dia_fim = f"//select[@id='content_content_content_Status_cboAberturaSecaoFimDia']/option[@value='{dia_fim}']"
    xpath_mes_fim = f"//select[@id='content_content_content_Status_cboAberturaSecaoFimMes']/option[@value='{mes_fim}']"
    xpath_ano_fim = f"//select[@id='content_content_content_Status_cboAberturaSecaoFimAno']/option[@value='{ano_fim}']"

    WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, xpath_dia_fim))).click()
    WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, xpath_mes_fim))).click()
    WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, xpath_ano_fim))).click()

    sleep(3)
    buscar = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.XPATH, "//input[@onclick='return verify();']"))
    )
    buscar.click()
    sleep(2)

    # Página inicial
    pagina_inicial_url = driver.current_url
    print('Página inicial:', pagina_inicial_url)

    cnpj_count = 0

    # Loop processando páginas
    while True:
        print("Extraindo dados da página atual...")
        proxima_pagina = captura_link(driver, ws, wb, pagina_inicial_url,cnpj_count, cnpj_label, janela)

        if not proxima_pagina:
            break  # Se não há mais próxima página, encerra o loop

    wb.save("dados_vencedores_diario_sp.xlsx")
    print("Planilha salva como 'dados_vencedores_diario_sp.xlsx'")
    driver.quit()
    print("Navegador fechado.")

def captura_link(driver, ws, wb, pagina_inicial_url,cnpj_count, cnpj_label, janela):
    links = driver.find_elements(By.XPATH, "//a[contains(@id, 'ResultadoBusca_dtgResultadoBusca_hlkObjeto')]")
    links_unicos = list(dict.fromkeys([link.get_attribute('href') for link in links]))

    print(f"Total de links encontrados nesta página: {len(links_unicos)}")

    for i, link in enumerate(links_unicos, start=1):
        print(f"Acessando Link {i}: {link}")
        driver.get(link)
        sleep(2)
        try:
            homologacao_links = driver.find_elements(By.XPATH, "//a[contains(@onclick, 'HOMOLOGAÇÃO')]")
            print(f"Total de links encontrados: {len(homologacao_links)}")
        except NoSuchElementException as e:
            print(f"Elemento não encontrado: {e}")
            homologacao_links = []
            
        sleep(2)
        if homologacao_links:
            homologacao_link = homologacao_links[0]
            onclick_content = homologacao_link.get_attribute('onclick')
            match = re.search(r"AbreJanelaDetalhes\((\d+),(\d+),\"HOMOLOGAÇÃO\"\)", onclick_content)
            if match:
                id_licitacao, id_evento = match.groups()
                homologacao_url = f"https://www.imprensaoficial.com.br/ENegocios/popup/pop_e-nego_detalhes.aspx?IdLicitacao={id_licitacao}&IdEventoLicitacao={id_evento}"
                print(f"Abrindo link de HOMOLOGAÇÃO: {homologacao_url}")
                try:
                    driver.get(homologacao_url)
                    print(f"URL atual: {driver.current_url}")
                except StaleElementReferenceException:
                    print("Elemento ficou stale. Tentando recapturar...")
                    driver.get(homologacao_url)
                
                sleep(2)

                try:
                    detalhes_texto = driver.find_element(By.ID, "content_content_content_DetalheEvento_lblSintesePublicacao").text
                    cnpj_match = re.search(r'CNPJ\s*(?:Nº|N.º|:|-)?\s*(\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2})', detalhes_texto)

                    if cnpj_match:  # Se houver uma correspondência válida
                        cnpj = cnpj_match.group(1)  # Obtém o CNPJ
                        cnpj_count += 1  # Incrementa a contagem
                    else:
                        cnpj = "Não encontrado"  # Se não encontrar, atribui "Não encontrado"

                    # Atualiza o label com a contagem de CNPJs extraídos
                    cnpj_label.configure(text=f"CNPJs extraídos: {cnpj_count}")
                    janela.update()


                    razao_social_match = re.search(r'(?:EMPRESA VENCEDORA:|a favor da empresa|EMPRESA\s*[:-]\s*)(.*?)(?=\s*CNPJ)', detalhes_texto)
                    razao_social = razao_social_match.group(1).strip() if razao_social_match else "Não encontrado"

                    ws.append([homologacao_url, cnpj, razao_social])
                    wb.save("dados_vencedores_diario_sp.xlsx")

                    
            

                except StaleElementReferenceException:
                    print("Elemento ficou stale. Tentando recapturar...")
                    detalhes_texto = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "content_content_content_DetalheEvento_lblSintesePublicacao"))
                    ).text
    else:
        print('Nao tem link de homologação')
    # Voltar à página inicial
    driver.get(pagina_inicial_url)
    sleep(2)

    # Tentar clicar em "Próxima" ao final da página
    try:
        botao_proxima = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//a[@id='content_content_content_ResultadoBusca_PaginadorCima_btnProxima']"))
        )
        botao_proxima.click()
        sleep(2)
        return True  # Retorna True para indicar que há próxima página
    except Exception as e:
        print(f"Não há mais próxima página: {e}")
        return False  # Retorna False quando não há mais páginas
    

def criar_interface_diario_sp():
    def capturar_datas():
        dia_inicio = combo_dia_inicio.get()
        mes_inicio = combo_mes_inicio.get()
        ano_inicio = combo_ano_inicio.get()

        dia_fim = combo_dia_fim.get()
        mes_fim = combo_mes_fim.get()
        ano_fim = combo_ano_fim.get()

        # Chama a função Selenium para preencher os dados em uma thread separada
        threading.Thread(target=iniciar_raspagem_compras_gov, args=(dia_inicio, mes_inicio, ano_inicio, dia_fim, mes_fim, ano_fim, cnpj_label, janela)).start()
    
    # Configuração inicial da janela principal usando customtkinter
    ctk.set_appearance_mode("Dark")  # Modo escuro
    ctk.set_default_color_theme("blue")  # Tema de cor azul

    janela = ctk.CTk()  # Usando CTk ao invés de Tk
    janela.title("Seleção de Datas")
    janela.geometry("500x400")

    # Componentes para Data Inicial
    label_inicio = ctk.CTkLabel(janela, text="Data Inicial:", font=("Arial", 14))
    label_inicio.pack(pady=10)

    frame_inicio = ctk.CTkFrame(janela)
    frame_inicio.pack(pady=5)

    combo_dia_inicio = ctk.CTkComboBox(frame_inicio, values=[str(i) for i in range(1, 32)], width=50)
    combo_dia_inicio.set("1")
    combo_dia_inicio.pack(side=tk.LEFT, padx=5)

    combo_mes_inicio = ctk.CTkComboBox(frame_inicio, values=[str(i) for i in range(1, 13)], width=50)
    combo_mes_inicio.set("1")
    combo_mes_inicio.pack(side=tk.LEFT, padx=5)

    combo_ano_inicio = ctk.CTkComboBox(frame_inicio, values=[str(i) for i in range(2004, 2026)], width=70)
    combo_ano_inicio.set("2024")
    combo_ano_inicio.pack(side=tk.LEFT, padx=5)

    # Componentes para Data Final
    label_fim = ctk.CTkLabel(janela, text="Data Final:", font=("Arial", 14))
    label_fim.pack(pady=10)

    frame_fim = ctk.CTkFrame(janela)
    frame_fim.pack(pady=5)

    combo_dia_fim = ctk.CTkComboBox(frame_fim, values=[str(i) for i in range(1, 32)], width=50)
    combo_dia_fim.set("1")
    combo_dia_fim.pack(side=tk.LEFT, padx=5)

    combo_mes_fim = ctk.CTkComboBox(frame_fim, values=[str(i) for i in range(1, 13)], width=50)
    combo_mes_fim.set("1")
    combo_mes_fim.pack(side=tk.LEFT, padx=5)

    combo_ano_fim = ctk.CTkComboBox(frame_fim, values=[str(i) for i in range(2004, 2026)], width=70)
    combo_ano_fim.set("2024")
    combo_ano_fim.pack(side=tk.LEFT, padx=5)

    # Botão para capturar as datas e preencher no Selenium
    btn_confirmar = ctk.CTkButton(janela, text="Iniciar Raspagem", command=capturar_datas, font=("Arial", 16))
    btn_confirmar.pack(pady=20)

    # Label para mostrar a quantidade de CNPJs extraídos
    cnpj_label = ctk.CTkLabel(janela, text="CNPJs extraídos: 0", font=("Arial", 14))
    cnpj_label.pack(pady=5)

    # Inicia o loop da interface
    janela.mainloop()


import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import customtkinter as ctk
from tkinter import messagebox
import os
import threading
from datetime import datetime
import logging

logging.basicConfig(level=logging.ERROR, format='%(asctime)s - %(levelname)s - %(message)s')


data_mais_recente = None
data_mais_antiga = None
intervalo_label = None

def atualizar_intervalo_datas(datas):
    global data_mais_recente, data_mais_antiga
    if datas:
        data_atual_mais_recente = max(datas)
        data_atual_mais_antiga = min(datas)
        
        # Atualizar as variáveis globais
        if not data_mais_recente or data_atual_mais_recente > data_mais_recente:
            data_mais_recente = data_atual_mais_recente
        if not data_mais_antiga or data_atual_mais_antiga < data_mais_antiga:
            data_mais_antiga = data_atual_mais_antiga

        # Atualizar o Label na interface gráfica
        intervalo_label.configure(
            text=f"Filtrado de {data_mais_antiga.strftime('%d/%m/%Y')} até {data_mais_recente.strftime('%d/%m/%Y')}"
        )

def extrair_datas_da_pagina(driver):
    datas_extradas = []
    try:
        elementos = driver.find_elements(By.XPATH, "//div[@class='datatable-body-cell-label']")
        for elemento in elementos:
            texto = elemento.text.strip()
            try:
                # Tentar converter o texto em uma data
                data = datetime.strptime(texto, "%d/%m/%Y")
                datas_extradas.append(data)
            except ValueError:
                # Ignorar valores que não sejam datas
                continue
    except Exception as e:
        logging.error(f"Erro ao extrair datas: {e}")
    return datas_extradas

def extrair_data_assinatura(driver):
    try:
        strong_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//strong[text()='Data de assinatura:']"))
        )

        # Encontrar o <span> imediatamente após o <strong> encontrado
        span_data_element = strong_element.find_element(By.XPATH, "following-sibling::span")

        # Extrair o texto da data
        data_texto = span_data_element.text.strip()

        # Retornar a data extraída
        return data_texto
    except Exception as e:
        logging.error(f"Erro ao extrair a data de assinatura: {e}")
        return "Não encontrada"

# Configurar o tema dark
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("dark-blue")

cnpj_count = 0

# Lista de sites de licitação/contratos
sites_licitacao = ["Portal Nacional de Contratações Públicas", "Outros Sites de Licitação"]


# Função para realizar a raspagem do Portal Nacional de Contratações Públicas
def iniciar_raspagem_Portal_Nacional_de_Contratacoes_Publicas():
    global cnpj_count
    driver = webdriver.Chrome()

    # Criar a planilha para salvar os dados
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Link", "CNPJ", "Razão Social", "Data de Divulgação no PNCP", "Data de Assinatura"])

    for pagina in range(1, 1000):  # De 1 a 999
        url = f'https://pncp.gov.br/app/contratos?q=&pagina={pagina}'
        driver.get(url)

        # Aguardar o carregamento dos elementos
        time.sleep(3)  # Espera de 3 segundos para garantir que a página carregue

        try:
            # Encontrar os botões únicos
            botoes = driver.find_elements(By.XPATH, "//a[contains(@class, 'br-item') and contains(@title, 'Acessar item.')]")
            botoes_unicos = list(dict.fromkeys([botao.get_attribute('href') for botao in botoes]))  # Remove duplicatas
            
            print(f"Página {pagina}: Total de botões encontrados: {len(botoes_unicos)}")

            for i, link in enumerate(botoes_unicos, start=1):
                try:
                    driver.get(link)
                    time.sleep(2)  # Espera um pouco para a página carregar

                    try:
                        # Extrair o CNPJ
                        cnpj_element = WebDriverWait(driver, 5).until(
                            EC.presence_of_element_located((By.XPATH, "//strong[contains(text(), 'CNPJ/CPF:')]/following-sibling::span"))
                        )
                        cnpj = cnpj_element.text
                        print(f"Botão {i}: CNPJ encontrado - {cnpj}")
                        cnpj_count += 1
                        cnpj_label.configure(text=f"CNPJs extraídos: {cnpj_count}")
                    except Exception as e:
                        print(f"Botão {i}: Erro ao extrair o CNPJ - {e}")
                        cnpj = "Não encontrado"

                    try:
                        # Extrair a Razão Social
                        razao_social_element = WebDriverWait(driver, 5).until(
                            EC.presence_of_element_located((By.XPATH, "//strong[contains(text(), 'Nome/Razão social:')]/following-sibling::span"))
                        )
                        razao_social = razao_social_element.text
                        print(f"Botão {i}: Razão social encontrada - {razao_social}")
                    except Exception as e:
                        print(f"Botão {i}: Erro ao extrair a Razão Social - {e}")
                        razao_social = "Não encontrado"

                    datas_pagina = extrair_datas_da_pagina(driver)
                    if datas_pagina:
                        data_divulgacao = datas_pagina[0].strftime("%d/%m/%Y")  # Usando a primeira data extraída
                    else:
                        data_divulgacao = "Não encontrada"
                    
                    data_texto = extrair_data_assinatura(driver)

                    # Atualizar intervalo de datas
                    atualizar_intervalo_datas(datas_pagina)

                    print(f"Data mais recente: {data_mais_recente.strftime('%d/%m/%Y')}")
                    print(f"Data mais antiga: {data_mais_antiga.strftime('%d/%m/%Y')}")
                    
                    
                    # Salvar os dados na planilha
                    ws.append([link, cnpj, razao_social, data_divulgacao, data_texto])
                    wb.save("dados_fornecedores.xlsx")

                except Exception as e:
                    print(f"Erro ao processar o botão {i}: {e}")
          
        except Exception as e:
            print(f"Erro ao acessar a página {pagina}: {e}")

    # Salvar a planilha
    
    driver.quit()
    messagebox.showinfo("Concluído", "Raspagem concluída com sucesso!")

def iniciar_raspagem():
    threading.Thread(target=iniciar_raspagem_Portal_Nacional_de_Contratacoes_Publicas, daemon=True).start()

# Criando a interface gráfica com CustomTkinter
def criar_interface_pncp():
    global cnpj_label
    global intervalo_label

    # Configuração da janela
    janela = ctk.CTk()
    janela.title("Raspagem de Dados de Licitação")
    janela.geometry("500x300")

    # Título
    titulo = ctk.CTkLabel(janela, text="Raspagem PNCP", font=("Arial", 18))
    titulo.pack(pady=20)

    # Botão para iniciar a raspagem
    btn_iniciar = ctk.CTkButton(janela, text="Iniciar Raspagem", command=iniciar_raspagem, font=("Arial", 16))
    btn_iniciar.pack(pady=20)

    # Label para mostrar o número de CNPJs extraídos
    cnpj_label = ctk.CTkLabel(janela, text="CNPJs extraídos: 0", font=("Arial", 14))
    cnpj_label.pack(pady=10)

    intervalo_label = ctk.CTkLabel(janela, text="Filtrado de --/--/---- até --/--/----", font=("Arial", 14))
    intervalo_label.pack(pady=10)

    # Iniciar a interface
    janela.mainloop()





def criar_interface_raspagem_emails():
    janela = ctk.CTk()  # Janela principal
    janela.geometry("500x300")
    janela.title("Raspagem de E-mails")

    # Label para exibir o número de e-mails raspados
    label_contador = ctk.CTkLabel(janela, text="E-mails raspados: 0", font=("Arial", 16))
    label_contador.pack(pady=10)

    # Combobox para selecionar o site
    label_selecione = ctk.CTkLabel(janela, text="Selecione o site:", font=("Arial", 14))
    label_selecione.pack(pady=5)

    site_selecionado = ctk.StringVar(value="CNPJ Já")
    dropdown_site = ctk.CTkComboBox(janela, values=["Consulta Guru", "CNPJ Já"], variable=site_selecionado)
    dropdown_site.pack(pady=5)

    # Função de depuração para verificar a seleção
    def verificar_selecao():
        print(f"Site selecionado: {site_selecionado.get()}")

    # Botão para iniciar a raspagem
    def iniciar_raspagem():
        site = site_selecionado.get()
        print(f"Valor do site selecionado no botão: {site}")
        if site == "Consulta Guru":
            consulta_cnpj_gratis(label_contador, janela)
        elif site == "CNPJ Já":
            consulta_cnpj_ja(label_contador, janela)

    # Botão de depuração para verificar o valor da seleção
    botao_verificar = ctk.CTkButton(janela, text="Verificar Seleção", command=verificar_selecao)
    botao_verificar.pack(pady=5)

    botao_iniciar = ctk.CTkButton(janela, text="Iniciar Raspagem", command=iniciar_raspagem)
    botao_iniciar.pack(pady=20)

    janela.mainloop()

def salvar_emails(resultados):
    """
    Salva os resultados no arquivo 'emails_vencedores.xlsx', evitando duplicatas.
    """
    df_resultados = pd.DataFrame(resultados)
    if not os.path.exists("emails_vencedores.xlsx"):
        df_resultados.to_excel("emails_vencedores.xlsx", index=False)
    else:
        df_existente = pd.read_excel("emails_vencedores.xlsx")
        df_atualizado = pd.concat([df_existente, df_resultados]).drop_duplicates(subset="Email", keep="first")
        df_atualizado.to_excel("emails_vencedores.xlsx", index=False)
    print("E-mails salvos no arquivo 'emails_vencedores.xlsx'.")

def consulta_cnpj_gratis(label_contador, janela):
    arquivos = [
        "dados_fornecedores.xlsx",
        "dados_vencedores_diario_sp.xlsx",
        "dados_vencedores_adjudicados_bll_compras.xlsx",
        "dados_vencedores_homologados_bll_compras.xlsx",
        "dados_fornecedores_5.xlsx"
    ]

    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

    def limpar_cnpj(cnpj):
        return ''.join(filter(str.isdigit, str(cnpj)))  # Mantém apenas dígitos

    total_emails = 0
    resultados = []
    emails_unicos = set()

    try:
        for arquivo in arquivos:
            if os.path.exists(arquivo):
                df = pd.read_excel(arquivo)

                for cnpj in df.iloc[:, 1]:  # Segunda coluna
                    cnpj_limpo = limpar_cnpj(cnpj)
                    if cnpj_limpo:
                        link = f"https://consulta.guru/consultar-cnpj-gratis/{cnpj_limpo}"
                        driver.get(link)
                        sleep(4)

                        try:
                            email_element = driver.find_element(By.XPATH, '//p[contains(text(), "@")]')
                            email = email_element.text.strip()

                            if email not in emails_unicos:
                                emails_unicos.add(email)
                                resultados.append({"CNPJ": cnpj_limpo, "Email": email})
                                total_emails += 1
                                label_contador.configure(text=f"E-mails raspados: {total_emails}")
                                janela.update()
                        except Exception:
                            pass  # Se não achar o email, passa para o próximo CNPJ

                        if total_emails % 5 == 0:
                            salvar_emails(resultados)
                            janela.after(60000, salvar_emails, resultados)  # 60 segundos após salvar

        salvar_emails(resultados)
    finally:
        driver.quit()

def consulta_cnpj_ja(label_contador, janela):
    arquivos = [
        "dados_fornecedores.xlsx",
        "dados_vencedores_diario_sp.xlsx",
        "dados_vencedores_adjudicados_bll_compras.xlsx",
        "dados_vencedores_homologados_bll_compras.xlsx",
        "dados_fornecedores_5.xlsx"
    ]

    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

    def limpar_cnpj(cnpj):
        return ''.join(filter(str.isdigit, str(cnpj)))

    total_emails = 0
    resultados = []
    emails_unicos = set()

    try:
        for arquivo in arquivos:
            if os.path.exists(arquivo):
                df = pd.read_excel(arquivo)

                for cnpj in df.iloc[:, 1]:
                    cnpj_limpo = limpar_cnpj(cnpj)
                    if cnpj_limpo:
                        link = f"https://cnpja.com/office/{cnpj_limpo}"
                        driver.get(link)
                        sleep(2)

                        try:
                            email_element = WebDriverWait(driver, 10).until(
                                EC.presence_of_element_located((By.XPATH, "//span[contains(text(), '@')]"))
                            )
                            email = email_element.text.strip()

                            if email not in emails_unicos:
                                emails_unicos.add(email)
                                resultados.append({"CNPJ": cnpj_limpo, "Email": email})
                                total_emails += 1
                                label_contador.configure(text=f"E-mails raspados: {total_emails}")
                                janela.update()
                        except Exception:
                            pass  # Se não achar o email, passa para o próximo CNPJ

                        if total_emails % 5 == 0:
                            salvar_emails(resultados)
                            janela.after(60000, salvar_emails, resultados)  # 60 segundos após salvar

        salvar_emails(resultados)
    finally:
        driver.quit()





def iniciar_raspagem():
    threading.Thread(target=iniciar_raspagem_Portal_Nacional_de_Contratacoes_Publicas, daemon=True).start()




# Configuração do tema escuro
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("dark-blue")

# Criação da janela principal
app = ctk.CTk()
app.title("Sistema de Raspagem de Licitações")
screen_width = app.winfo_screenwidth()
screen_height = app.winfo_screenheight()
app.geometry(f"{screen_width}x{screen_height-40}+0+0")

# Título principal
title_label = ctk.CTkLabel(
    app, 
    text="Sistema de Raspagem de Licitações", 
    font=("Helvetica", 30, "bold"), 
    text_color="#ffffff"
)
title_label.pack(pady=20)

# Frame principal que contém todos os módulos
main_frame = ctk.CTkFrame(app, fg_color="transparent")
main_frame.pack(padx=20, pady=20, fill="both", expand=True)

# Frame para os 5 módulos
frame_modulos = ctk.CTkFrame(main_frame)
frame_modulos.pack(pady=30, fill="both", expand=True)

# Módulo 1 - Portal Nacional de Licitações
frame_pncp = ctk.CTkFrame(frame_modulos)
frame_pncp.pack(side="left", padx=20, pady=20, fill="y", expand=True)

# Imagem de exemplo para o módulo
img_pncp = ctk.CTkImage(Image.open("img.png"), size=(300, 100))  # Substitua pelo caminho correto da imagem
label_img_pncp = ctk.CTkLabel(frame_pncp, image=img_pncp, text="")
label_img_pncp.pack(pady=10)

label_pncp = ctk.CTkLabel(frame_pncp, text="Portal Nacional de Contratações Públicas", font=("Helvetica", 18, "bold"))
label_pncp.pack(pady=10)

button_pncp = ctk.CTkButton(
    frame_pncp, 
    text="Iniciar Raspagem", 
    command=criar_interface_pncp,
    font=("Helvetica", 14), 
    width=250, 
    height=40, 
    fg_color="#007acc",  
    hover_color="#005b99",  
)
button_pncp.pack(pady=10)

button_enviar_pncp = ctk.CTkButton(
    frame_pncp, 
    text="Enviar E-mails", 
    command=enviar_emails_site1,
    font=("Helvetica", 14), 
    width=250, 
    height=40, 
    fg_color="#4caf50",  
    hover_color="#388e3c",  
)
button_enviar_pncp.pack(pady=10)

# Módulo 2 - BLL - COMPRAS
frame_site1 = ctk.CTkFrame(frame_modulos)
frame_site1.pack(side="left", padx=20, pady=20, fill="y", expand=True)

# Imagem de exemplo para o módulo
img_site1 = ctk.CTkImage(Image.open("bll_compras.png"), size=(300, 100))  # Substitua pelo caminho correto da imagem
label_img_site1 = ctk.CTkLabel(frame_site1, image=img_site1, text="")
label_img_site1.pack(pady=10)

label_site1 = ctk.CTkLabel(frame_site1, text="BLL Compras", font=("Helvetica", 18, "bold"))
label_site1.pack(pady=10)

button_site1 = ctk.CTkButton(
    frame_site1, 
    text="Iniciar Raspagem", 
    command=criar_interface,
    font=("Helvetica", 14), 
    width=250, 
    height=40, 
    fg_color="#007acc",  
    hover_color="#005b99",  
)
button_site1.pack(pady=10)

button_enviar_site1 = ctk.CTkButton(
    frame_site1, 
    text="Enviar E-mails", 
    command=enviar_emails_site2,
    font=("Helvetica", 14), 
    width=250, 
    height=40, 
    fg_color="#4caf50",  
    hover_color="#388e3c",  
)
button_enviar_site1.pack(pady=10)

# Módulo 3 - Outro Site de Licitações
frame_site2 = ctk.CTkFrame(frame_modulos)
frame_site2.pack(side="left", padx=20, pady=20, fill="y", expand=True)

# Imagem de exemplo para o módulo
img_site2 = ctk.CTkImage(Image.open("diario_sp.png"), size=(300, 100))  # Substitua pelo caminho correto da imagem
label_img_site2 = ctk.CTkLabel(frame_site2, image=img_site2, text="")
label_img_site2.pack(pady=10)

label_site2 = ctk.CTkLabel(frame_site2, text="Diário de São Paulo", font=("Helvetica", 18, "bold"))
label_site2.pack(pady=10)

button_site2 = ctk.CTkButton(
    frame_site2, 
    text="Iniciar Raspagem", 
    command=criar_interface_diario_sp,
    font=("Helvetica", 14), 
    width=250, 
    height=40, 
    fg_color="#007acc",  
    hover_color="#005b99",  
)
button_site2.pack(pady=10)

button_enviar_site2 = ctk.CTkButton(
    frame_site2, 
    text="Enviar E-mails", 
    command=enviar_emails_site3,
    font=("Helvetica", 14), 
    width=250, 
    height=40, 
    fg_color="#4caf50",  
    hover_color="#388e3c",  
)
button_enviar_site2.pack(pady=10)

# Rodapé
footer_frame = ctk.CTkFrame(app, fg_color="transparent")
footer_frame.pack(side="bottom", pady=20)

button_eliminar = ctk.CTkButton(
    footer_frame, 
    text="Eliminar CNPJ Repetido", 
    command=eliminar_cnpjs,
    font=("Helvetica", 14), 
    width=250, 
    height=40, 
    fg_color="#ff5722",  
    hover_color="#d43f00",  
)
button_eliminar.pack(pady=10)

button_raspagem_emails = ctk.CTkButton(
    footer_frame, 
    text="Raspagem de E-mails", 
    command=criar_interface_raspagem_emails,
    font=("Helvetica", 14), 
    width=250, 
    height=40, 
    fg_color="#007acc",  
    hover_color="#005b99",  
)
button_raspagem_emails.pack(side="left", padx=10)


# Iniciar o loop principal da interface gráfica
app.mainloop()
#app